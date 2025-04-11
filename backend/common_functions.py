import os
from dateutil import parser
from openpyxl.styles import Font
from pypdf import PdfReader
import re
import shutil
import io
import openpyxl
from openpyxl.styles import Alignment
import time
import pandas as pd
import numpy as np
import PyPDF2
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
from datetime import datetime, timedelta
import regex as re
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import datefinder
from calendar import monthrange
import calendar
from openpyxl.styles import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
import logging
import openpyxl
from openpyxl.styles import Alignment
from .utils import get_base_dir
import fitz
bold_font = Font(bold=True)
pd.options.display.float_format = "{:,.2f}".format
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)
logger = logging.getLogger(__name__)
BASE_DIR = get_base_dir()
logger.info("Base Dir : ", BASE_DIR)
#from old_bank_extractions import CustomStatement
import json
from .code_for_extraction import extract_text_from_pdf, extract_with_test_cases, model_for_pdf, extract_dataframe_from_pdf

##EXTRACTION PROCESS
def extract_text_from_file(file_path):

    file_extension = os.path.splitext(file_path)[1].lower()

    if file_extension == ".csv":
        df = pd.read_csv(file_path)
        text = df.head(20).to_string(index=False)

    else:
        df = pd.read_excel(file_path)
        text = df.head(20).to_string(index=False)

    return text

def add_start_n_end_date( df, start_date, end_date, bank):
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce")
    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")

    # Check if the period falls within the start and end dates
    start_date_sd = pd.to_datetime(start_date, format="%d-%m-%Y", errors="coerce")
    end_date_ed = pd.to_datetime(end_date, format="%d-%m-%Y", errors="coerce")
    period_start = pd.to_datetime(
        df["Value Date"].iloc[0], format="%d-%m-%Y", errors="coerce"
    )
    period_end = pd.to_datetime(
        df["Value Date"].iloc[-1], format="%d-%m-%Y", errors="coerce"
    )

    if (start_date_sd - timedelta(days=1)) <= period_start <= (
        end_date_ed + timedelta(days=1)
    ) and (start_date_sd - timedelta(days=1)) <= period_end <= (
        end_date_ed + timedelta(days=1)
    ):
        print("The period falls within the start and end dates.")
    else:
        raise Exception(
            f"Error: The period for Bank: {bank} ({period_start} to {period_end}), "
            f"does not fall within the start and end dates ({start_date_sd} to {end_date_ed}), provided by the user."
        )

    # add opening and closing balance
    start_bal = (
        df.iloc[0]["Balance"] - df.iloc[0]["Credit"]
        if df.iloc[0]["Credit"] > 0
        else df.iloc[0]["Balance"] + df.iloc[0]["Debit"]
    )
    end_bal = df.iloc[-1]["Balance"]

    start_row = pd.DataFrame(
        [
            {
                "Value Date": start_date,
                "Description": "Opening Balance",
                "Debit": 0,
                "Credit": 0,
                "Balance": start_bal,
            }
        ]
    )
    end_row = pd.DataFrame(
        [
            {
                "Value Date": end_date,
                "Description": "Closing Balance",
                "Debit": 0,
                "Credit": 0,
                "Balance": end_bal,
            }
        ]
    )

    idf = pd.concat([start_row, df, end_row], ignore_index=True)
    idf["Bank"] = f"{bank}"
    return idf


# Function to extract account number and IFSC code from text
def extract_info(raw_text):
    acc = "XXXXXXXXXXXXXX"
    # Regular expressions to capture Customer ID and ECS No
    customer_id_pattern = r"(?i)\bCust(?:omer)?\s*ID[:\s]*\d{10}\b"
    ecs_no_pattern = r"(?i)ECS\s*No[:\s]*\d{10,18}\b"  # Match ECS No followed by digits
    cif_no_pattern = r"CIF\sNo\.?\s*[:\-]?\s*(\d+)"

    # Remove Customer ID and ECS No from the raw text
    if re.search(customer_id_pattern, raw_text):
        raw_text = re.sub(customer_id_pattern, "", raw_text)
    if re.search(ecs_no_pattern, raw_text):
        raw_text = re.sub(ecs_no_pattern, "", raw_text)
    if re.search(cif_no_pattern, raw_text):
        raw_text = re.sub(cif_no_pattern, "", raw_text)

    # Patterns specifically targeting the numeric part for account numbers
    account_number_patterns = [
        r"\b(?!18002026161\b)(?!9\d{9}\b)(?!91\d{10}\b)\d{10,18}\b",
        r"(?!CIF No\.?:?\s*\d+\s*)Account No\.?\s*[:#]?\s*(\d+\/?[A-Z]*\/?\d+)",
        r"(?i)CUSTOMER\s*ID[:\s]*\d{10}\b",
        r"(?!Phone No. :?\s?)\d(10)",
        r"Account No\s*[:#]?\s*(\d+)",
        r"Account\s*No\s*[:#]?\s*(\d+)",
        r"Account\sNo.\s:\s(\d+\/[A-Z]+\/\d+)",
        r"Account\sNo\.?\s*[:#]?\s*(\d+\/[A-Z]+\/\d+)",
        r"Account\s*number\s*[:#]?\s*(\d+)",
        r"account\s*number\s*[:#]?\s*(\d+)",
        r"Account\s*Number\s*[:#]?\s*(\d+)",
        r"Account Number\s*:\s*(\d+)",
        r"Account Number\s*(\d+)",
        r"A/C NO[:#]?\s*(\d+)",
        r"STATEMENT PERIOD\s*(\d+)",
        r"Account number[:#]?\s*(\d+)",
        r"Account\s*[:#]?\s*(\d+)",
        r"\b\d{15}\b",
        r"Account #\s*(\d+)",
        r"\b\d{3}-\d{6}-\d{3}\b",
        r"A/c X{10}\d{4}",
        r"\b\d{8}\b",
    ]

    ifsc_pattern = r"\b[A-Z]{4}0[A-Z0-9]{6}\b"  # IFSC code pattern

    if not raw_text:  # If raw_text is None or empty, return None values
        return acc

    # Search for account number using specific patterns
    for pattern in account_number_patterns:
        match = re.search(pattern, raw_text, re.IGNORECASE)

        if match:
            try:
                # Try accessing group 1 if it exists
                acc = match.group(1).strip()  # Extracted number only
            except IndexError:
                # If group 1 does not exist, return the whole match
                acc = match.group(0).strip()  # Whole match
        else:
            acc = None  # No match found

        if match:
            return acc


    # Return None if no pattern matches
    return acc

def extract_account_details(text):

    try:
        # Combined pattern to match account holder names for different banks
        name_patterns = [
            re.compile(p, re.IGNORECASE)
            for p in [
                r"Customer Details\s*:\s*(.*?)\s*\n",
                r"Name\s*:\s*([^\n]+)",
                r"Account Holders? Name\s*:\s*([^\n]+)",
                r"(?:MR\.?|M/S\.?|MS\.?|MRS\.?)\s*([^\n]+)",
                r"Name:\s*(.*)",
                r"date ofstatement\s*\n(.*)",
                r"(.*)\s*Period",
                r"Customer Name\s*:\s*(.*)",
                r"INR\s*\n(.*)",
                r"CUSTOMER NAME (.*)",
                r"Customer\s*Details\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"Account\s*Holder\s*Name\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"(?:MR\.?|MRS\.?|MS\.?|M/S\.?)\s*([A-Z][a-zA-Z\s]*[A-Z])",
                r"Customer\s*Name\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"Name\s*of\s*Customer\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"Name\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"Accountholder\s*Name\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])\s*",
                r"Account\s*Title\s*[:\-]?\s*([A-Z][A-Z\s]+[A-Z])",
                r"CUSTOMER\s*NAME\s*[:\-]?\s*([A-Z][a-zA-Z\s]+[A-Z])",
                r"To\s*,?\s*([A-Z][a-zA-Z\s]+[A-Z])",
                r"TO\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])",
                r"Account\s*Title\s*:\s*([A-Z][a-zA-Z\s]+[A-Z])",
                r"Name\s+([A-Z][A-Z\s]+[A-Z])",
                r"Account Holders? Name\s*([A-Z][A-Z\s]+[A-Z])(?:\s*\n)?",
            ]
        ]

        names = []
        for pattern in name_patterns:
            matches = pattern.findall(text)
            if matches:
                names.extend(matches)
        # Fallback for joint holder text specific to AXIS_BANK
        joint_holder_text = "Joint Holder :"
        if joint_holder_text in text:
            parts = text.split(joint_holder_text, 1)
            names.extend(parts[0].strip().split("\n"))

        names = [
            name.strip() for name in names if name.strip()
        ]  # Clean up names list

        # Combined pattern to match account numbers for different banks
        account_numbers = extract_info(text)

        details = [
            names[0] if names else "_____",
            account_numbers,
        ]

        return details

    except Exception as e:
        print(
            f"An error occurred while extracting names and account numbers: {str(e)}"
        )
        return ["_", "XXXXXXXXXX"]

def extract_extension( path):
    # Get the file extension
    print("Path:", path)
    _, extension = os.path.splitext(path)
    extension = extension.lower()

    # Dictionary mapping file extensions to pandas read functions
    # read_functions = {
    #     '.xlsx': pd.read_excel,
    #     '.xls': pd.read_excel,
    #     '.xlsm': pd.read_excel,
    #     '.xlsb': pd.read_excel,
    #     '.csv': pd.read_csv,
    #     '.xltx': pd.read_excel,
    #     '.xltm': pd.read_excel
    # }

    # # Check if the extension is supported
    # if extension not in read_functions:
    #     raise ValueError(f"Unsupported file extension: {extension}")

    return extension

import csv
def convert_csv_to_excel(csv_path, CA_ID):
    # Read the CSV with Python's built-in CSV reader to avoid parsing errors
    os.makedirs("saved_csv", exist_ok=True)
    excel_path = os.path.join("saved_csv", f"temp_{CA_ID}_excel.xlsx")

    with open(csv_path, 'r', encoding='utf-8') as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    # Save the content as an Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    wb.save(excel_path)

    return excel_path


def extraction_process(bank, pdf_path, pdf_password, start_date, end_date):
    CA_ID = "1234_temp"
    empty_idf = pd.DataFrame()
    default_name_n_num = ["_", "XXXXXXXXXX"]
    a = ""
    # bank = re.sub(r"\d+", "", bank)
    ext = extract_extension(pdf_path)

    try:
        if ext == ".pdf":
            idf, text, explicit_lines = extract_with_test_cases(bank, pdf_path, pdf_password, CA_ID)
            name_n_num = explicit_lines if idf.empty else extract_account_details(text)

        elif ext == ".csv":
            pdf_path = convert_csv_to_excel(pdf_path, CA_ID)
            df = pd.read_excel(pdf_path)
            df.loc[0] = df.columns
            df.columns = range(df.shape[1])

            start_index = df.apply(
                lambda row: (
                    row.astype(str).str.contains("date", case=False).any() and
                    row.astype(str).str.contains("balance|total amount", case=False).any()) or
                    row.astype(str).str.contains("balance|total amount", case=False).any(),
                axis=1
            ).idxmax()
            df = df.loc[start_index:] if start_index is not None else pd.DataFrame()
            idf, _ = model_for_pdf(df)
            name_n_num = extract_account_details(extract_text_from_file(pdf_path))

        else:
            df = pd.read_excel(pdf_path)
            df.loc[0] = df.columns
            df.columns = range(df.shape[1])

            start_index = df.apply(
                lambda row: (
                    row.astype(str).str.contains("date", case=False).any() and
                    row.astype(str).str.contains("balance|total amount", case=False).any()) or
                    row.astype(str).str.contains("balance|total amount", case=False).any(),
                axis=1
            ).idxmax()
            df = df.loc[start_index:] if start_index is not None else pd.DataFrame()
            idf, _ = model_for_pdf(df)
            name_n_num = extract_account_details(extract_text_from_file(pdf_path))

        if not idf.empty:
            idf = add_start_n_end_date(idf, start_date, end_date, bank)

        return idf, name_n_num, a

    except Exception as e:
        return empty_idf, default_name_n_num, str(e)


def extraction_process_explicit_lines(bank, pdf_path, pdf_password, start_date, end_date, explicit_lines, labels):
    CA_ID = "1234_temp"
    empty_idf = pd.DataFrame()
    default_name_n_num = ["_", "XXXXXXXXXX"]
    # bank = re.sub(r"\d+", "", bank)
    a = ""

    try:
        df = extract_dataframe_from_pdf(pdf_path, table_settings={
            "vertical_strategy": "explicit",
            "explicit_vertical_lines": explicit_lines,
            "horizontal_strategy": "lines",
            "intersection_x_tolerance": 120,
        })

        all_null = all(label[1] == "null" for label in labels)

        if not all_null:
            new_row = [None] * len(df.columns)  # Create a blank row with the same number of columns
            for index, label_type in labels:
                if index < len(new_row):
                    new_row[index] = label_type

            # Insert the new row at the top of the DataFrame
            df.loc[-1] = new_row  # Add the new row with a negative index to place it at the top
            df.index = df.index + 1  # Shift all indices by 1
            df.sort_index(inplace=True)  # Reorder the DataFrame to update the row positions

        idf, _ = model_for_pdf(df)
        name_n_num = extract_account_details(extract_text_from_pdf(pdf_path))

        # Add start and end date
        if idf.empty:
            df = extract_dataframe_from_pdf(pdf_path, table_settings={
                "vertical_strategy": "explicit",
                "explicit_vertical_lines": explicit_lines,
                "horizontal_strategy": "text",
            })

            all_null = all(label[1] == "null" for label in labels)

            if not all_null:
                new_row = [None] * len(df.columns)  # Create a blank row with the same number of columns
                for index, label_type in labels:
                    if index < len(new_row):
                        new_row[index] = label_type

                # Insert the new row at the top of the DataFrame
                df.loc[-1] = new_row  # Add the new row with a negative index to place it at the top
                df.index = df.index + 1  # Shift all indices by 1
                df.sort_index(inplace=True)  # Reorder the DataFrame to update the row positions

            idf, _ = model_for_pdf(df)
            name_n_num = extract_account_details(extract_text_from_pdf(pdf_path))

        idf = add_start_n_end_date(idf, start_date, end_date, bank)

        return idf, name_n_num, a

    except Exception as e:
        
        df = extract_dataframe_from_pdf(pdf_path, table_settings={
            "vertical_strategy": "explicit",
            "explicit_vertical_lines": explicit_lines,
            "horizontal_strategy": "text",
            "intersection_x_tolerance": 120,
        })

        all_null = all(label[1] == "null" for label in labels)

        if not all_null:
            new_row = [None] * len(df.columns)  # Create a blank row with the same number of columns
            for index, label_type in labels:
                if index < len(new_row):
                    new_row[index] = label_type

            # Insert the new row at the top of the DataFrame
            df.loc[-1] = new_row  # Add the new row with a negative index to place it at the top
            df.index = df.index + 1  # Shift all indices by 1
            df.sort_index(inplace=True)  # Reorder the DataFrame to update the row positions

        idf, _ = model_for_pdf(df)
        name_n_num = extract_account_details(extract_text_from_pdf(pdf_path))

        # Add start and end date
        if not idf.empty:
            idf = add_start_n_end_date(idf, start_date, end_date, bank)
            return idf, name_n_num, a
        else:
            return empty_idf, default_name_n_num, str(e)

##EOD
def monthly( df):
    # add a new row with the average of month values in each column
    e = df.copy()
    e.replace(0, np.nan, inplace=True)
    new_row = pd.DataFrame(e.iloc[0:31].mean(axis=0)).T.round(2)
    monthly_avg = pd.concat([df, new_row], ignore_index=True)
    monthly_avg.iloc[-1, 0] = "Average"
    return monthly_avg

def eod(df_original):
    df = df_original.copy()
    df["Value Date"] = pd.to_datetime(
        df["Value Date"], format="%d-%m-%Y", errors="coerce")
    df.dropna(subset=["Value Date"], inplace=True)
    if df.empty: return pd.DataFrame()  # Handle empty df after NaT drop
    df["Balance"] = pd.to_numeric(df["Balance"], errors="coerce")
    df.dropna(subset=["Balance"], inplace=True)  # Option: Drop rows with invalid balances
    if df.empty: return pd.DataFrame()  # Handle empty df after balance cleaning
    df.sort_values(by=["Bank", "Value Date"], inplace=True)  # Sort by Bank too for consistency
    df["YearMonth"] = df["Value Date"].dt.strftime("%Y%m")  # Keep as string initially for easier date range generation
    df["MonthStr"] = df["Value Date"].dt.strftime("%b-%Y")
    df["Date"] = df["Value Date"].dt.day
    global_max_date = df["Value Date"].max() if not df.empty else None
    df = df[["Value Date", "Balance", "YearMonth", "MonthStr", "Date", "Bank"]]
    bank_names = df["Bank"].unique().tolist()
    multiple_eods = []
    all_banks_processed_months = set()  # Keep track of all month columns across banks
    for bank in bank_names:
        idf = df[df["Bank"] == bank].copy()  # Use copy for safety
        if idf.empty: continue  # Skip banks with no valid data
        min_ym_str = idf['YearMonth'].min()
        max_ym_str = idf['YearMonth'].max()
        start_date = pd.to_datetime(min_ym_str, format='%Y%m')
        end_date = pd.to_datetime(max_ym_str, format='%Y%m')
        all_month_starts = pd.date_range(start=start_date, end=end_date, freq='MS')
        result_eod_bank = pd.DataFrame()  # Accumulate results for this bank
        previous_month_end_balance = None  # Reset for each bank
        processed_months_strs_bank = []  # Keep track of month strings in order for THIS bank
        for month_start in all_month_starts:
            current_ym_str = month_start.strftime('%Y%m')
            current_month_str = month_start.strftime('%b-%Y')
            processed_months_strs_bank.append(current_month_str)  # Add to ordered list for pivoting
            all_banks_processed_months.add(current_month_str)  # Add to global set for final merging
            month_data_exists = current_ym_str in idf['YearMonth'].values
            eod_month_df = pd.DataFrame()  # Initialize DataFrame for the current month
            if month_data_exists:
                month_subset = idf.loc[idf["YearMonth"] == current_ym_str]
                eod_month_df = month_subset.drop_duplicates(subset="Date", keep="last").copy()  # Use copy
                eod_month_df = eod_month_df.set_index("Date")
                all_days_index = pd.RangeIndex(start=1, stop=32, name="Date")
                eod_month_df = eod_month_df.reindex(all_days_index)
                eod_month_df["Bank"] = bank
                eod_month_df["MonthStr"] = current_month_str
                eod_month_df["YearMonth"] = current_ym_str
                eod_month_df.reset_index(inplace=True)  # Get 'Date' back as column
                if previous_month_end_balance is not None and pd.isna(eod_month_df.loc[0, "Balance"]):
                    first_valid_idx = eod_month_df['Balance'].first_valid_index()
                    if first_valid_idx is not None:
                        eod_month_df.loc[0:first_valid_idx - 1, 'Balance'] = previous_month_end_balance
                    else:  # Whole month was NaN initially (only possible if input month_subset was empty, but check exists)
                        eod_month_df['Balance'] = previous_month_end_balance
                eod_month_df["Balance"] = eod_month_df["Balance"].ffill()
                fill_value = 0.0 if previous_month_end_balance is None else previous_month_end_balance
                eod_month_df["Balance"] = eod_month_df["Balance"].fillna(fill_value)
                if not eod_month_df.empty:
                    previous_month_end_balance = eod_month_df["Balance"].iloc[-1]

            else:
                fill_balance = 0.0 if previous_month_end_balance is None else previous_month_end_balance
                eod_month_df = pd.DataFrame({
                    "Date": range(1, 32),
                    "Balance": fill_balance,
                    "YearMonth": current_ym_str,
                    "MonthStr": current_month_str,
                    "Bank": bank,
                    "Value Date": pd.NaT  # Or None
                })
            result_eod_bank = pd.concat([result_eod_bank, eod_month_df], ignore_index=True)
        if result_eod_bank.empty: continue  # Skip if bank had no processable data at all
        pivot_df = result_eod_bank.pivot(
            index="Date", columns="MonthStr", values="Balance"
        )  # Index is 'Date' (1-31)
        pivot_df = pivot_df.reindex(columns=processed_months_strs_bank)
        pivot_df.reset_index(inplace=True)
        pivot_df.rename(columns={"Date": "Day"}, inplace=True)
        day_col_name = "Day"  # Use the actual column name
        pivot_numeric_cols = pivot_df.columns.difference([day_col_name])  # Exclude 'Day'/'Date'
        mask = pd.DataFrame(False, index=pivot_df.index, columns=pivot_numeric_cols)
        for col in pivot_numeric_cols:
            try:
                month_dt = pd.to_datetime(col, format="%b-%Y")
                days_in_month = month_dt.days_in_month
                # Use the correct day column name here
                mask[col] = pivot_df[day_col_name] > days_in_month
            except ValueError:
                print(f"Could not parse column '{col}' to determine days in month during zeroing.",
                                UserWarning)
        pivot_df.loc[:, pivot_numeric_cols] = pivot_df.loc[:, pivot_numeric_cols].mask(mask, 0.0)
        if global_max_date and not pivot_df.empty:
            last_month_in_pivot = pivot_df.columns[-1]
            try:
                last_month_dt = pd.to_datetime(last_month_in_pivot, format="%b-%Y")
                global_max_month_dt = pd.Timestamp(global_max_date.year, global_max_date.month, 1)
                if last_month_dt == global_max_month_dt:
                    end_day_of_global_max = global_max_date.day
                    pivot_df.loc[pivot_df[day_col_name] > end_day_of_global_max, last_month_in_pivot] = 0.0
            except ValueError:
                print(f"Could not parse last column '{last_month_in_pivot}' during final day zeroing.",
                                UserWarning)
        multiple_eods.append(pivot_df)
    if not multiple_eods:
        return pd.DataFrame()  # No data processed for any bank
    if len(multiple_eods) == 1:
        adf = multiple_eods[0]
    else:
        adf = process_repeating_columns(
            multiple_eods)  # Make sure this function correctly merges based on 'Day'
    final_ordered_months = sorted(
        list(all_banks_processed_months),
        key=lambda x: pd.to_datetime(x, format='%b-%Y', errors='coerce')
    )
    day_col_name = "Day" if "Day" in adf.columns else ("Date" if "Date" in adf.columns else None)
    if day_col_name:
        adf = adf.reindex(columns=[day_col_name] + final_ordered_months, fill_value=0.0)
    else:
        adf = adf.reindex(columns=final_ordered_months, fill_value=0.0)
    day_col_name = "Day" if "Day" in adf.columns else None  # Find Day column name again
    numeric_cols_for_sum = adf.columns.difference([day_col_name]) if day_col_name else adf.columns
    sum_row_data = adf.iloc[0:31][numeric_cols_for_sum].sum(axis=0)
    if day_col_name:
        sum_row_data[day_col_name] = 'Total'  # Add the label for the 'Day' column
    total_row_df = pd.DataFrame([sum_row_data], columns=adf.columns)
    total_df = pd.concat([adf, total_row_df], ignore_index=True)
    all_df = monthly(total_df)  # Assuming this function exists
    return all_df

def opening_and_closing_bal(edf): 
        import warnings
        opening_bal = {}
        closing_bal = {}
        month_columns = [col for col in edf.columns if col != 'Day']
        if not month_columns or edf.empty:
            warnings.warn("Input DataFrame 'edf' is empty or has no valid month columns.", UserWarning)
            return opening_bal, closing_bal
        edf_data_only = edf.copy()
        if edf_data_only.iloc[-1, 0] == "Total":
            edf_data_only = edf_data_only.iloc[:-1]
            edf_data_only.reset_index(drop=True, inplace=True)
        for month_str in month_columns:
            try:
                month_dt = pd.to_datetime(month_str, format="%b-%Y")
                days_in_month = month_dt.days_in_month
                last_day_index = days_in_month - 1
                if 0 <= last_day_index < len(edf_data_only):
                    balance = edf_data_only.iloc[last_day_index][month_str]
                    closing_bal[month_str] = float(balance)
                else:
                    closing_bal[month_str] = np.nan
            except ValueError:
                closing_bal[month_str] = np.nan
            except KeyError:
                warnings.warn(f"Column '{month_str}' not found during closing balance calculation.", UserWarning)
                closing_bal[month_str] = np.nan
            except Exception as e:
                warnings.warn(f"An unexpected error occurred calculating closing balance for {month_str}: {e}", UserWarning)
                closing_bal[month_str] = np.nan
        ordered_months = month_columns
        for i, month in enumerate(ordered_months):
            if i == 0:
                warnings.warn(f"Using balance from Day 1 of month '{month}' in the input DataFrame "
                              "as the Opening Balance for this first month. Ensure this value "
                              "represents the start-of-day balance.", UserWarning)
                try:
                    first_day_balance_val = edf_data_only.iloc[0][month]
                    opening_bal[month] = float(first_day_balance_val)
                except (KeyError, IndexError):
                    warnings.warn(f"Could not retrieve balance from input DataFrame for Day 1 of the first month ({month}). "
                                  "Setting opening balance to NaN.", UserWarning)
                    opening_bal[month] = np.nan
                except (ValueError, TypeError):
                    warnings.warn(f"Balance value ('{first_day_balance_val}') from input DataFrame for Day 1 "
                                  f"of the first month ({month}) could not be converted to float. Setting opening balance to NaN.",
                                   UserWarning)
                    opening_bal[month] = np.nan
            else:
                prev_month = ordered_months[i - 1]
                if prev_month in closing_bal and not pd.isna(closing_bal[prev_month]):
                    opening_bal[month] = closing_bal[prev_month]
                else:
                    warnings.warn(f"Could not find a valid closing balance for the previous month ({prev_month}) "
                                  f"to use as opening balance for {month}. Setting opening balance to NaN.", UserWarning)
                    opening_bal[month] = np.nan

        return opening_bal, closing_bal
    
def avgs_df( df):
    # quarterly_avg
    if df.shape[1] > 3:
        df_chi_list_1 = []
        # Iterate through every three columns in the original DataFrame
        for i in range(1, df.shape[1], 3):
            # Get the current three columns
            subset = df.iloc[:, i : i + 3]
            if subset.shape[1] < 3:
                new_row = 0.0
            else:
                new_row = subset.iloc[-2].sum() / 3
            subset.loc[len(subset)] = new_row
            df_chi_list_1.append(subset)
        result_df = pd.concat(df_chi_list_1, axis=1)
        new_row = pd.Series({"Day": "Quarterly Average"})
        df = df._append(new_row, ignore_index=True)
        result_df.insert(0, "Day", df["Day"])
        df = result_df

        # half - yearly avg
        if df.shape[1] > 6:
            df_chi_list_2 = []
            # Iterate through every three columns in the original DataFrame
            for i in range(1, df.shape[1], 6):
                # Get the current three columns
                subset = df.iloc[:, i : i + 6]
                if subset.shape[1] < 6:
                    new_row = 0.0
                else:
                    new_row = subset.iloc[-3].sum() / 6
                subset.loc[len(subset)] = new_row
                df_chi_list_2.append(subset)
            result_df = pd.concat(df_chi_list_2, axis=1)
            new_row = pd.Series({"Day": "Half-Yearly Average"})
            df = df._append(new_row, ignore_index=True)
            result_df.insert(0, "Day", df["Day"])
            df = result_df

            # yearly avg
            if df.shape[1] > 12:
                df_chi_list_3 = []
                # Iterate through every three columns in the original DataFrame
                for i in range(1, df.shape[1], 12):
                    # Get the current three columns
                    subset = df.iloc[:, i : i + 12]
                    if subset.shape[1] < 12:
                        new_row = 0.0
                    else:
                        new_row = subset.iloc[-4].sum() / 12
                    subset.loc[len(subset)] = new_row
                    df_chi_list_3.append(subset)
                result_df = pd.concat(df_chi_list_3, axis=1)
                new_row = pd.Series({"Day": "Yearly Average"})
                df = df._append(new_row, ignore_index=True)
                result_df.insert(0, "Day", df["Day"])
                df = result_df

            else:
                new_row = pd.Series({"Day": "Yearly Average"})
                df = df._append(new_row, ignore_index=True)

        else:
            new_row = pd.Series({"Day": "Half-Yearly Average"})
            df = df._append(new_row, ignore_index=True)

    else:
        new_row = pd.Series({"Day": "Quarterly Average"})
        df = df._append(new_row, ignore_index=True)

    return df

def calculate_fixed_day_average(data):
        day_col_name = 'Day'
        if day_col_name not in data.columns:
            return pd.DataFrame()  # Return empty if essential column is missing
        df = data.copy()
        month_columns = [col for col in df.columns if col != day_col_name]
        if not month_columns:
            return pd.DataFrame()
        calc_df = df[pd.to_numeric(df[day_col_name], errors='coerce').notna()]
        calc_df[day_col_name] = calc_df[day_col_name].astype(int)  # Ensure Day is int
        calc_df = calc_df[calc_df[day_col_name].between(1, 31)].reset_index(drop=True)
        for col in month_columns:
            calc_df[col] = pd.to_numeric(calc_df[col], errors='coerce').fillna(
                0)  # Coerce errors to NaN, then fill with 0
        daily_averages = {day_col_name: "Daily_Avg"}
        for month_str in month_columns:
            try:
                month_dt = pd.to_datetime(month_str, format="%b-%Y")
                days_in_month = month_dt.days_in_month
                valid_day_data = calc_df.loc[calc_df[day_col_name] <= days_in_month, month_str]
                month_sum = valid_day_data.sum()
                if days_in_month > 0:
                    daily_averages[month_str] = month_sum / float(days_in_month)
                else:
                    daily_averages[month_str] = np.nan
            except Exception as e:
                daily_averages[month_str] = np.nan
        Average_df = pd.DataFrame([daily_averages])
        sets_of_days = [
            {"days": [5, 15, 25], "label": "Avg_Days_5_15_25"},
            {"days": [5, 10, 15, 25], "label": "Avg_Days_5_10_15_25"},
            {"days": [1, 5, 10, 15, 20, 25], "label": "Avg_Days_1_5_10_15_20_25"},
            {"days": [8, 10, 15, 20, 25], "label": "Avg_Days_8_10_15_20_25"},
            {"days": [1, 5, 10, 15, 20], "label": "Avg_Days_1_5_10_15_20"},
            {"days": [5, 10, 15, 20, 25], "label": "Avg_Days_5_10_15_20_25"},
            {"days": [1, 7, 14, 21, 28], "label": "Avg_Days_1_7_14_21_28"},
            {"days": [1, 5, 10, 15, 25], "label": "Avg_Days_1_5_10_15_25"},
            {"days": [5, 15, 25, 30], "label": "Avg_Days_5_15_25_30"},
            {"days": [2, 4, 10, 17, 21], "label": "Avg_Days_2_4_10_17_25"},
            {"days": [5, 15, 25, 30], "label": "Avg_Days_5_15_25_30"},
            {"days": [1, 5, 15, 20, 25], "label": "Avg_Days_1_5_15_20_25"},
            {"days": [4, 5, 7, 10, 15, 25], "label": "Avg_Days_4_5_7_10_15_25"},
            {"days": [5, 10, 15, 20, 25, 30], "label": "Avg_Days_5_10_15_20_25_30"},
            {"days": [5, 10, 15, 20, 26], "label": "Avg_Days_5_10_15_20_26"},
            {"days": [1, 5, 10, 18, 25], "label": "Avg_Days_1_5_10_18_25"},
            {"days": [2, 10, 20, 30], "label": "Avg_Days_2_10_20_30"},
        ]
        avg_balance_df_list = []
        for day_set in sets_of_days:
            selected_days_subset = calc_df[calc_df[day_col_name].isin(day_set["days"])]
            set_averages = {day_col_name: day_set["label"]}
            for month_str in month_columns:
                try:
                    month_dt = pd.to_datetime(month_str, format="%b-%Y")
                    days_in_month = month_dt.days_in_month
                    valid_selected_days = selected_days_subset[selected_days_subset[day_col_name] <= days_in_month]
                    month_avg = valid_selected_days[month_str].mean()
                    set_averages[month_str] = month_avg
                except Exception as e:
                    set_averages[month_str] = np.nan
            average_balance_df = pd.DataFrame([set_averages])
            avg_balance_df_list.append(average_balance_df)
        all_avg_balances = pd.concat([Average_df] + avg_balance_df_list, ignore_index=True)
        numeric_cols = all_avg_balances.columns.difference([day_col_name])
        all_avg_balances[numeric_cols] = all_avg_balances[numeric_cols].round(2)
        averages_with_monthly = calculate_monthly_averages(all_avg_balances)
        return averages_with_monthly

def process_avg_last_6_months(data, eod):
    def manish(loan_value_df):
        if (
                loan_value_df.empty
                or "Maximum Home Loan Value" not in loan_value_df.columns
                or "Maximum LAP Value" not in loan_value_df.columns
                or "Maximum BL Value" not in loan_value_df.columns
        ):
            print("DataFrame is empty or required columns are missing.")
            return  # or handle this situation in a way that fits your application

        # Your existing logic with safety checks
        max_home_loan = (
            0
            if pd.isna(loan_value_df["Maximum Home Loan Value"].iloc[0])
            else round(loan_value_df["Maximum Home Loan Value"].iloc[0] / 1000) * 1000
        )
        max_lap = (
            0
            if pd.isna(loan_value_df["Maximum LAP Value"].iloc[0])
            else round(loan_value_df["Maximum LAP Value"].iloc[0] / 1000) * 1000
        )
        max_bl = (
            0
            if pd.isna(loan_value_df["Maximum BL Value"].iloc[0])
            else round(loan_value_df["Maximum BL Value"].iloc[0] / 1000) * 1000
        )

        commission_percentage = [
            0.45 / 100,
            0.65 / 100,
            1.00 / 100,
        ]  # Convert percentages to fractions

        # Calculate commission in Rs
        commission_home_loan = round(max_home_loan * commission_percentage[0], 2)
        commission_lap = round(max_lap * commission_percentage[1], 2)
        commission_bl = round(max_bl * commission_percentage[2], 2)

        # Create the dataframe
        data = {
            "Product": [
                "Home Loan / Balance Transfer",
                "Loan Against Property / Balance Transfer",
                "Business Loan",
                "Term Plan",
                "General Insurance",
            ],
            "Amount": [max_home_loan, max_lap, max_bl, np.nan, np.nan],
            "Commission %": ["0.45%", "0.65%", "1.00%", "1%-30%", "upto 10%"],
            "Commission (in Rs)": [
                commission_home_loan,
                commission_lap,
                commission_bl,
                np.nan,
                np.nan,
            ],
        }
        df = pd.DataFrame(data)
        return df

    error_df = pd.DataFrame()

    # Check if 'Avg_Last_6_Months' column exists
    if "Avg_Last_6_Months" not in data.columns:
        print("'Avg_Last_6_Months' column not found in DataFrame.")
        return error_df  # Or return an empty DataFrame as needed

    # Extract the second value from 'Avg_Last_6_Months' and perform calculations
    if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 1:
        if not pd.isna(data["Avg_Last_6_Months"].iloc[1]):
            avg_divided_by_1_5 = data["Avg_Last_6_Months"].iloc[1] / 1.5
        else:
            avg_divided_by_1_5 = np.nan
    else:
        print(
            "'Avg_Last_6_Months' column does not exist or does not have enough data."
        )
        avg_divided_by_1_5 = np.nan

    # For 'Avg_Last_12_Months' at index 0
    if "Avg_Last_12_Months" in data.columns:
        if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
            avg_divided_by_2_idfc = data["Avg_Last_12_Months"].iloc[0] / 2
        else:
            avg_divided_by_2_idfc = np.nan
    else:
        print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
        avg_divided_by_2_idfc = np.nan

    # For 'Avg_Last_6_Months' at index 2
    if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 2:
        if not pd.isna(data["Avg_Last_6_Months"].iloc[2]):
            avg_divided_by_2_indus = data["Avg_Last_6_Months"].iloc[2] / 1.5
        else:
            avg_divided_by_2_indus = np.nan
    else:
        print(
            "'Avg_Last_6_Months' column does not exist or does not have enough data."
        )
        avg_divided_by_2_indus = np.nan

    # For 'Avg_Last_12_Months' at index 0 again
    if "Avg_Last_12_Months" in data.columns:
        if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
            avg_divided_by_2_L_T = data["Avg_Last_12_Months"].iloc[0] / 2
        else:
            avg_divided_by_2_L_T = np.nan
    else:
        print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
        avg_divided_by_2_L_T = np.nan

    annual_interest_rate = 0.0870
    periods = 20 * 12
    principal = 100000
    payment_value = pmt(principal, annual_interest_rate, periods)
    payment_for_lap = pmt_lap()
    payment_for_bl = pmt_bl()

    # Calculating Loan value for axis
    axis_home_loan_value = None
    if payment_value != 0:
        axis_home_loan_value = avg_divided_by_1_5 / payment_value
        axis_home_loan_value = axis_home_loan_value * 100000
        axis_home_loan_value = round(axis_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    axis_LAP_value = None
    if payment_for_lap != 0:
        axis_LAP_value = avg_divided_by_1_5 / payment_for_lap
        axis_LAP_value = axis_LAP_value * 100000
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    axis_bl_value = None
    if payment_for_bl != 0:
        axis_bl_value = avg_divided_by_1_5 / payment_for_bl
        axis_bl_value = axis_bl_value / payment_for_lap
        axis_bl_value = axis_bl_value * 100000
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    # Calculating loan value for Idfc
    Idfc_home_loan_value = None
    if payment_value != 0:
        Idfc_home_loan_value = avg_divided_by_2_idfc / payment_value
        Idfc_home_loan_value = Idfc_home_loan_value * 100000
        Idfc_home_loan_value = round(Idfc_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    Idfc_LAP_value = None
    if payment_for_lap != 0:
        Idfc_LAP_value = avg_divided_by_2_idfc / payment_for_lap
        Idfc_LAP_value = Idfc_LAP_value * 100000
        Idfc_LAP_value = round(Idfc_LAP_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    Idfc_bl_value = None
    if payment_for_bl != 0:
        Idfc_bl_value = avg_divided_by_2_idfc / payment_for_bl
        Idfc_bl_value = Idfc_bl_value * 100000
        Idfc_bl_value = round(Idfc_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    # Calculating loan value for Indus
    indus_home_loan_value = None
    if payment_value != 0:
        indus_home_loan_value = avg_divided_by_2_indus / payment_value
        indus_home_loan_value = indus_home_loan_value * 100000
        indus_home_loan_value = round(indus_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    indus_LAP_value = None
    if payment_for_lap != 0:
        indus_LAP_value = avg_divided_by_2_indus / payment_for_lap
        indus_LAP_value = indus_LAP_value * 100000
        indus_LAP_value = round(indus_LAP_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    indus_bl_value = None
    if payment_for_bl != 0:
        indus_bl_value = avg_divided_by_2_indus / payment_for_bl
        indus_bl_value = indus_bl_value * 100000
        indus_bl_value = round(indus_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    L_T_home_loan_value = None
    if payment_value != 0:
        L_T_home_loan_value = avg_divided_by_2_L_T / payment_value
        L_T_home_loan_value = L_T_home_loan_value * 100000
        L_T_home_loan_value = round(L_T_home_loan_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")
    L_T_LAP_value = None
    if payment_for_lap != 0:
        L_T_LAP_value = avg_divided_by_2_L_T / payment_for_lap
        L_T_LAP_value = L_T_LAP_value * 100000
        L_T_LAP_value = round(L_T_LAP_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    L_T_bl_value = None
    if payment_for_lap != 0:
        L_T_bl_value = avg_divided_by_2_L_T / payment_for_bl
        L_T_bl_value = L_T_bl_value * 100000
        L_T_bl_value = round(L_T_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    home_loan_values = []

    # Append home loan values to the list
    if axis_home_loan_value is not None:
        home_loan_values.append(axis_home_loan_value)

    if Idfc_home_loan_value is not None:
        home_loan_values.append(Idfc_home_loan_value)

    if indus_home_loan_value is not None:
        home_loan_values.append(indus_home_loan_value)

    if L_T_home_loan_value is not None:
        home_loan_values.append(L_T_home_loan_value)

    # Calculate the maximum home loan value
    max_home_loan_value = (
        max(home_loan_values, default=None) if home_loan_values else None
    )

    lap_values = []
    bl_values = []

    if axis_LAP_value is not None:
        lap_values.append(axis_LAP_value)

    if Idfc_LAP_value is not None:
        lap_values.append(Idfc_LAP_value)

    if indus_LAP_value is not None:
        lap_values.append(indus_LAP_value)

    if L_T_LAP_value is not None:
        lap_values.append(L_T_LAP_value)

    if axis_bl_value is not None:
        bl_values.append(axis_bl_value)

    if Idfc_bl_value is not None:
        bl_values.append(Idfc_bl_value)

    if indus_bl_value is not None:
        bl_values.append(indus_bl_value)

    if L_T_bl_value is not None:
        bl_values.append(L_T_bl_value)

    # Calculate the maximum values
    max_lap_value = max(lap_values, default=None) if lap_values else None
    max_bl_value = max(bl_values, default=None) if bl_values else None

    # Create DataFrame with maximum values
    max_values_df = pd.DataFrame(
        {
            "Maximum Home Loan Value": [max_home_loan_value],
            "Maximum LAP Value": [max_lap_value],
            "Maximum BL Value": [max_bl_value],
        }
    )

    new_df = manish(max_values_df)
    return new_df

def process_repeating_columns( oy):
    df = pd.concat(oy, axis=1)
    df = df.loc[:, ~df.columns.duplicated(keep="first") | (df.columns != "Day")]
    repeating_columns = [
        col for col in df.columns if df.columns.tolist().count(col) > 1
    ]

    idf = pd.DataFrame(
        {col: df[col].sum(axis=1).round(4) for col in repeating_columns}
    )
    df = df.drop(columns=repeating_columns)
    concatenated_df = pd.concat([df, idf], axis=1)

    sorted_columns = sorted(
        [col for col in concatenated_df.columns if col != "Day"],
        key=lambda x: pd.to_datetime(x, format="%b-%Y"),
    )
    sorted_columns_formatted = [
        col.strftime("%b-%Y") if isinstance(col, pd.Timestamp) else col
        for col in sorted_columns
    ]
    concatenated_df = concatenated_df[["Day"] + sorted_columns_formatted]
    return concatenated_df

def calculate_monthly_averages( data):
    original_columns = data.columns.tolist()

    # Calculate the average for the last 12 months
    if len(original_columns) >= 12:
        last_12_months = original_columns[-12:]
        if data[last_12_months].applymap(np.isreal).all().all():
            data["Avg_Last_12_Months"] = data[last_12_months].mean(axis=1).round(2)
        else:
            print(
                "Non-numeric data found in last 12 months columns, 'Avg_Last_12_Months' will not be added."
            )

    # Update the columns list to exclude the new average column
    updated_columns = data.columns.tolist()

    # Calculate the average for the last 6 months
    if len(original_columns) >= 6:
        last_6_months = original_columns[-6:]
        if data[last_6_months].applymap(np.isreal).all().all():
            data["Avg_Last_6_Months"] = data[last_6_months].mean(axis=1).round(2)
        else:
            print(
                "Non-numeric data found in last 6 months columns, 'Avg_Last_6_Months' will not be added."
            )

    # Similarly, update the columns list again to exclude the new average column
    updated_columns = data.columns.tolist()

    # Calculate the average for the last 18 months
    if len(original_columns) >= 18:
        last_18_months = original_columns[-18:]
        if data[last_18_months].applymap(np.isreal).all().all():
            data["Avg_Last_18_Months"] = data[last_18_months].mean(axis=1).round(2)
        else:
            print(
                "Non-numeric data found in last 18 months columns, 'Avg_Last_18_Months' will not be added."
            )
    return data

def pmt( principal, annual_interest_rate, periods):
    r = annual_interest_rate / 12  # monthly interest rate
    n = periods
    payment = principal * r / (1 - (1 + r) ** -n)
    # print("pmt", payment)
    return payment

def pmt_lap():
    annual_interest_rate = 0.0950
    periods = 15 * 12
    principal = 100000
    r = annual_interest_rate / 12  # monthly interest rate
    n = periods
    payment = principal * r / (1 - (1 + r) ** -n)
    # print("pmt_lap", payment)
    return payment

def pmt_bl():
    annual_interest_rate = 0.2200
    periods = 3 * 12
    principal = 100000
    r = annual_interest_rate / 12  # monthly interest rate
    n = periods
    payment = principal * r / (1 - (1 + r) ** -n)
    # print("pmt_bl", payment)
    return payment


def category_add_ca(df):
    x = df["Balance"]
    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].str.lower()
    df["Description"] = df["Description"].str.replace(" ", "")
    excel_file_path = os.path.join(BASE_DIR, "Final_Category.xlsx")
    excel2 = os.path.join(BASE_DIR, "Customer_category.xlsx")
    df1 = pd.read_excel(excel_file_path)
    df2_additional = pd.read_excel(excel2)
    df2 = pd.concat([df1, df2_additional], ignore_index=True)
    print("excel_file_path -",excel_file_path)
    # df2 = pd.read_excel(excel_file_path)

    # Initialize the 'Category' column with "Suspense" for all rows
    df["Category"] = "Suspense"

    pos_pattern = r"^pos.*"
    df.loc[(df["Description"].str.contains(pos_pattern, regex=True)) & (df["Debit"] > 0), "Category",] = "POS-Dr"
    df.loc[(df["Description"].str.contains(pos_pattern, regex=True)) & (df["Credit"] > 0), "Category",] = "POS-Cr"

    pos_pattern_2 = r"^(vps|ips|ecom|pur|pcd|edc|ecompur)"
    df.loc[(df["Description"].str.contains(pos_pattern_2, regex=True)) & (df["Debit"] > 0), "Category",] = "POS-Dr"
    df.loc[(df["Description"].str.contains(pos_pattern_2, regex=True)) & (df["Credit"] > 0), "Category",] = "POS-Cr"

    def categorize_bank_charges(df):
        Bank_charges = r"^(bctt|nchg|tChg|tip/scg|rate\.diff|owchquereturncharges|inwardchqreturncharge|chrg|incidentalcharges|iwchq|smschrg|chrg:sms|\*chrg:sms|nachreturncharges|fundtransfercharges|cashwithdrawalchgs|impschg|monthlysmscha|amcatmcharges|monthlyservicechrgs|smsalert|penalcharges|sgst|cgst|bulkcharges)"
        df = df[~df["Description"].str.contains("POS-Cr|POS-Dr", regex=True, na=False)]
        df.loc[(df["Description"].str.contains(Bank_charges, regex=True)) & (
                    df["Debit"] > 0), "Category",] = "Bank Charges"
        return df

    df = categorize_bank_charges(df)

    Bank_charges = r"(wchrgs)"
    df.loc[(df["Description"].str.contains(Bank_charges, case=False, regex=True)) & (
                df["Debit"] > 0), "Category"] = "Bank Charges"

    Bank_Interest_Recieved = r"^(int)"
    df.loc[(df["Description"].str.contains(Bank_Interest_Recieved, regex=True)) & (
                df["Credit"] > 0), "Category",] = "Bank Interest Received"

    Bounce = r"^(r-ret-utr)"
    df.loc[(df["Description"].str.contains(Bounce, regex=True)) & (df["Debit"] > 0), "Category",] = "Bounce"

    Cash_Withdrawal = r"^(ccwd|vat|mat|nfs|atm|atm-cash-axis|atm-cash|atw|csw|atd|ati|vmt|inf|cwdr|self|cash-atm|atl/|cashpm|withdrawal|chequewdl)"
    df.loc[(df["Description"].str.contains(Cash_Withdrawal, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Cash Withdrawal"

    General_insurance = r"^(pac)"
    df.loc[(df["Description"].str.contains(General_insurance, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "General insurance"

    Indirect_tax = r"^(idtx)"
    df.loc[(df["Description"].str.contains(Indirect_tax, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Indirect tax"

    interest_paid = r"^(int.coll)"
    df.loc[(df["Description"].str.contains(interest_paid, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "interest paid"

    investment = r"^(eba|autosweep|growwpay|axismutualfund)"
    df.loc[(df["Description"].str.contains(investment, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Investment"
    investment = r"(growwpay)"
    df.loc[(df["Description"].str.contains(investment, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Investment"

    Local_cheque_collection = r"^(lccbrncms)"
    df.loc[(df["Description"].str.contains(Local_cheque_collection, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Local cheque collection"

    emi = r"^(emi|lnpy)"
    df.loc[
        (df["Description"].str.contains(emi, case=False, regex=True)) & (df["Debit"] > 0), "Category",] = "Probable EMI"

    Tax_Payment = r"^(gib)"
    df.loc[(df["Description"].str.contains(Tax_Payment, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "Tax Payment"
    Tax_Payment = r"(gsttaxpayment|gst@)"
    df.loc[(df["Description"].str.contains(Tax_Payment, case=False, regex=True)) & (
                df["Debit"] > 0), "Category",] = "GST Paid"

    Refund_Reversal = r"^(ft-rev|revchrg|rev:imps|imps:rec|imps_ret)"
    df.loc[(df["Description"].str.contains(Refund_Reversal, case=False, regex=True)) & (
                df["Credit"] > 0), "Category",] = "Refund/Reversal"

    Refund = r"^(imps:rec|ref-tr)"
    df.loc[(df["Description"].str.contains(Refund, case=False, regex=True)) & (
                df["Credit"] > 0), "Category",] = "Refund/Reversal"

    Redemption = r"^(revsweep|sewwptrf)"
    df.loc[(df["Description"].str.contains(Redemption, case=False, regex=True)) & (
                df["Credit"] > 0), "Category",] = "Redemption,Dividend & Interest"

    Recharge = r"^(rchg)"
    df.loc[(df["Description"].str.contains(Recharge, case=False, regex=True)) & (
                df["Credit"] > 0), "Category",] = "Recharge"

    # Function to extract and clean name from 'ipay/inst/neft/' transactions
    def extract_ipay_neft_name(description):
        if "neft" in description.lower():
            try:
                name_part = description.split("neft", 1)[1]
                name_cleaned = re.sub(r"[/\d\s]+", " ", name_part).strip()
                name_match = re.search(r"[a-zA-Z]+", name_cleaned)
                if name_match:
                    return name_match.group(0).strip()  # Cleaned Name (without slashes/spaces)
                else:
                    return "Suspense"  # Default if name is not found
            except IndexError:
                return "Suspense"

    NEFT_ipay = df[df["Description"].str.contains("ipay/inst/neft/", case=False, na=False)].copy()
    if not NEFT_ipay.empty:
        NEFT_ipay["Category"] = NEFT_ipay["Description"].apply(extract_ipay_neft_name)
        df.update(NEFT_ipay)

    # SBI
    NEFT_SBI = df[df["Description"].str.contains("bytransfer-neft*", na=False)]
    if not NEFT_SBI.empty:
        NEFT_1 = NEFT_SBI[~NEFT_SBI["Category"].str.contains("Redemption, Dividend & Interest")]

        def extract_category(description):
            try:
                return description.split("*")[3]
            except IndexError:
                return "Suspense"

        neft_names = NEFT_1["Description"].apply(extract_category)
        NEFT_1["Category"] = neft_names
        df.update(NEFT_1)

    def extract_text_after_numeric(s):
        matches = re.findall(
            r"\d+([^\d]+)", s
        )  # Regex to find digits followed by non-digits
        return (
            matches[-1] if matches else None
        )

    def extract_neft_name(description):
        parts = description.split("/")
        if len(parts) > 2 and parts[2].strip():
            return parts[2].strip()
        else:
            return "Suspense"

    NEFT = df[df["Description"].str.contains("neft", na=False)]
    NEFT_1 = NEFT[NEFT["Description"].str.contains("neft/", na=False)]
    if not NEFT_1.empty:
        NEFT_1 = NEFT_1[NEFT_1["Category"].str.contains("Suspense", na=False)]
        exclude_pattern = "Redemption|Dividend & Interest|Salary Paid|Salary Received"
        NEFT_1 = NEFT_1[~NEFT_1["Category"].str.contains(exclude_pattern, na=False)]
        NEFT_1["Category"] = NEFT_1["Description"].apply(extract_neft_name)
        df.update(NEFT_1)

    NEFT_KOTAK = NEFT[~NEFT["Category"].str.contains(
        "Debtor|Creditor|Suspense|Redemption,Dividend & Interest,Salary Paid,Salary Received", na=False, )]
    if not NEFT_KOTAK.empty:
        extracted_text = NEFT_KOTAK["Description"].apply(extract_text_after_numeric)
        NEFT_KOTAK["Category"] = extracted_text
        df.update(NEFT_KOTAK)

    def extract_neft_colon_category(description):
        try:
            name_part = description.split("neft:")[1]
            name_without_numbers = re.sub(r"\d+$", "", name_part)
            return name_without_numbers.strip()
        except IndexError:
            return "Suspense"  # Or any other default value you prefer

    NEFT_colon = df[df["Description"].str.contains("neft:", na=False)]
    if not NEFT_colon.empty:
        NEFT_colon = NEFT_colon[~NEFT_colon["Category"].str.contains(
            "Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received", na=False, )]
        NEFT_colon["Category"] = NEFT_colon["Description"].apply(extract_neft_colon_category)
        df.update(NEFT_colon)

    def extract_nefto_union_category(description):
        try:
            name_part = description.split("nefto-")[1]
            name_without_numbers = re.sub(r"\d+$", "", name_part)
            return name_without_numbers.strip()
        except IndexError:
            return "Suspense"  # Or any other default value you prefer

    NEFTO = df[df["Description"].str.contains("nefto-", na=False)]
    if not NEFTO.empty:
        NEFTO = NEFTO[~NEFTO["Category"].str.contains(
            "Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received", na=False, )]
        NEFTO["Category"] = NEFTO["Description"].apply(extract_nefto_union_category)
        df.update(NEFTO)

    def extract_rtgso_union_category(description):
        try:
            name_part = description.split("rtgso-")[1]
            name_without_numbers = re.sub(r"\d+$", "", name_part)
            return name_without_numbers.strip()
        except IndexError:
            return "Suspense"

    RTGSO = df[df["Description"].str.contains("rtgso-", na=False)]
    if not RTGSO.empty:
        RTGSO = RTGSO[~RTGSO["Category"].str.contains(
            "Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received", na=False, )]
        RTGSO["Category"] = RTGSO["Description"].apply(extract_rtgso_union_category)
        df.update(RTGSO)

    def extract_rtgsfr_category(description):
        try:
            category_part = description.split("rtgsfr:")[1].split("/")[0]
            return category_part
        except IndexError:
            return "Suspense"

    RTGSFR = df[df["Description"].str.contains("rtgsfr:", na=False)]
    if not RTGSFR.empty:
        RTGSFR = RTGSFR[~RTGSFR["Category"].str.contains(
            "Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received", na=False, )]
        RTGSFR["Category"] = RTGSFR["Description"].apply(extract_rtgsfr_category)
        df.update(RTGSFR)

    def extract_rtgso_union_category(description):
        try:
            name_part = description.split("rtgs:")[1]
            name_without_numbers = re.sub(r"\d+$", "", name_part)
            return name_without_numbers.strip()
        except IndexError:
            return "Suspense"

    RTGSO = df[df["Description"].str.contains("rtgs:", na=False)]
    if not RTGSO.empty:
        RTGSO = RTGSO[~RTGSO["Category"].str.contains(
            "Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received", na=False, )]
        RTGSO["Category"] = RTGSO["Description"].apply(extract_rtgso_union_category)
        df.update(RTGSO)

    rtgs_ib = df[df["Description"].str.contains("ib/rtgs/", na=False)]
    rtgs_ib = rtgs_ib[
        ~rtgs_ib["Category"].str.contains("Debtor|Creditor|Redemption, Dividend & Interest|Salary Paid|Salary Received",
                                          na=False, )]
    if not rtgs_ib.empty:
        def extract_category(description):
            parts = description.split("/")
            return (
                parts[3] if len(parts) > 3 else "Suspense"
            )  # Default value for descriptions without enough parts

        rtgs_names = rtgs_ib["Description"].apply(extract_category)
        rtgs_ib["Category"] = rtgs_names
        df.update(rtgs_ib)

    NEFT_ib = df[df["Description"].str.contains("ib/neft/", na=False)]
    if not NEFT_ib.empty:
        NEFT_ib = NEFT_ib[
            ~NEFT_ib["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]

        def extract_category(description):
            parts = description.split("/")
            return (
                parts[3] if len(parts) > 3 else "Suspense"
            )  # Default value for descriptions without enough parts

        neft_names = NEFT_ib["Description"].apply(extract_category)
        NEFT_ib["Category"] = neft_names
        df.update(NEFT_ib)

    NEFT = df[df["Description"].str.contains("neft/mb/ax", na=False)]
    if not NEFT.empty:
        NEFT = NEFT[~NEFT["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]

        def extract_category(description):
            parts = description.split("/")
            return (
                parts[3] if len(parts) > 3 else "Suspense"
            )  # Default value for descriptions without enough parts

        neft_names = NEFT["Description"].apply(extract_category)
        NEFT["Category"] = neft_names
        df.update(NEFT)

    def extract_keyword_from_description(description):
        parts = description.split("--")
        if len(parts) < 2:
            return None
        words = parts[0].split("-")
        return max(words, key=len, default=None)

    NEFT_entries = df[df["Description"].str.contains("neft-", na=False)]
    NEFT_entries = NEFT_entries[
        ~NEFT_entries["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]
    if not NEFT_entries.empty:
        for idx, row in NEFT_entries.iterrows():
            keyword = extract_keyword_from_description(row["Description"])
            if keyword and not keyword.isdigit():
                df.at[idx, "Category"] = keyword
            else:
                df.at[idx, "Category"] = "Suspense"

    def extract_neft_io_category(description):
        try:
            category_part = description.split("-")[3]
            if any(char.isdigit() for char in category_part):
                return "Suspense"
            return category_part
        except IndexError:
            return "Suspense"

    NEFT_IO = df[df["Description"].str.contains("neft-", na=False) & df["Category"].str.contains("Suspense", na=False)]
    if not NEFT_IO.empty:
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains(
            "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Debtor")]
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Creditor")]
        NEFT_IO["Category"] = NEFT_IO["Description"].apply(extract_neft_io_category)
        df.update(NEFT_IO)

    def extract_neft_hdfc_cr_category(description):
        parts = description.split("-")
        return (
            parts[2] if len(parts) > 2 else "Suspense"
        )  # Default value for descriptions without enough parts

    NEFT_HDFC_CR = df[df["Description"].str.contains("neftcr", na=False)]
    if not NEFT_HDFC_CR.empty:
        NEFT_HDFC_CR = NEFT_HDFC_CR[~NEFT_HDFC_CR["Category"].str.contains(
            "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
        neft_names = NEFT_HDFC_CR["Description"].apply(extract_neft_hdfc_cr_category)
        NEFT_HDFC_CR["Category"] = neft_names
        df.update(NEFT_HDFC_CR)

    NEFT_HDFC_DR = df[df["Description"].str.contains("neftdr", na=False)]
    if not NEFT_HDFC_DR.empty:
        NEFT_HDFC_DR = NEFT_HDFC_DR[
            ~NEFT_HDFC_DR["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]

        def extract_category(description):
            try:
                return (
                    description.split("-")[2] if "-" in description else description
                )
            except IndexError:
                return "Suspense"

        neft_names = NEFT_HDFC_DR["Description"].apply(extract_category)
        NEFT_HDFC_DR["Category"] = neft_names
        df.update(NEFT_HDFC_DR)

    NEFT_thane = df[df["Description"].str.contains("toneft", na=False)]
    if not NEFT_thane.empty:
        NEFT_thane = NEFT_thane[
            ~NEFT_thane["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]

        def extract_category(description):
            try:
                return description.split("/")[1]
            except IndexError:
                return "Suspense"  # Default to 'Suspense' in case of IndexError

        neft_thane_names = NEFT_thane["Description"].apply(extract_category)
        NEFT_thane["Category"] = neft_thane_names
        df.update(NEFT_thane)

    NEFT_UCO = df[(df["Description"].str.contains("neft/", na=False)) & (df["Category"] == "Suspense")]
    if not NEFT_UCO.empty:
        NEFT_1 = NEFT_UCO[
            ~NEFT_UCO["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]
        NEFT_1 = NEFT_1[~NEFT_1["Category"].str.contains("Debtor")]
        NEFT_1 = NEFT_1[~NEFT_1["Category"].str.contains("Creditor")]

        def extract_category(description):
            try:
                return description.split("/")[1]
            except IndexError:
                return "Suspense"  # Default to 'Suspense' in case of IndexError

        neft_uco_names = NEFT_1["Description"].apply(extract_category)
        NEFT_1["Category"] = neft_uco_names
        df.update(NEFT_1)

    def extract_net_neft_category(description):
        try:
            # Split by 'rtgsfr:' and take the first part after it
            category_part = description.split("net/neft/")[1].split("/")[0]
            return category_part
        except IndexError:
            # In case of an IndexError, return 'Suspense'
            return "Suspense"

    net_neft = df[df["Description"].str.contains("net/neft/", na=False)]
    net_neft = net_neft[
        ~net_neft["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]
    if not net_neft.empty:
        net_neft["Category"] = net_neft["Description"].apply(extract_net_neft_category)
        df.update(net_neft)

    def extract_nft_category(description):
        try:
            category_part = description.split("nft/")[1].split("/")[0]
            return category_part
        except IndexError:
            return "Suspense"

    nft_neft = df[df["Description"].str.contains("nft/", na=False)]
    nft_neft = nft_neft[
        ~nft_neft["Category"].str.contains("Redemption, Dividend & Interest|Salary Paid|Salary Received")]
    if not nft_neft.empty:
        nft_neft["Category"] = nft_neft["Description"].apply(extract_nft_category)
        df.update(nft_neft)

    def extract_bob_neft_category(description):
        try:
            category_part = description.split("-")[2]
            if any(char.isdigit() for char in category_part):
                return "Suspense"
            return category_part
        except IndexError:
            return "Suspense"

    NEFT_BOB = df[df["Description"].str.contains("neft-", na=False)]
    if not NEFT_BOB.empty:
        NEFT_BOB = NEFT_BOB[~NEFT_BOB["Category"].str.contains(
            "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
        NEFT_BOB = NEFT_BOB[~NEFT_BOB["Category"].str.contains("Debtor")]
        NEFT_BOB = NEFT_BOB[~NEFT_BOB["Category"].str.contains("Creditor")]
        NEFT_BOB["Category"] = NEFT_BOB["Description"].apply(extract_bob_neft_category)
        df.update(NEFT_BOB)

    def extract_neft_category(description):
        try:
            name_part = description.split("-")[3]
            if any(char.isdigit() for char in name_part):
                return "Suspense"
            return name_part
        except IndexError:
            return "Suspense"

    NEFT_IO = df[df["Description"].str.contains("neft-", na=False) & df["Category"].str.contains("Suspense", na=False)]
    if not NEFT_IO.empty:
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains(
            "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Debtor")]
        NEFT_IO = NEFT_IO[~NEFT_IO["Category"].str.contains("Creditor")]
        NEFT_IO["Category"] = NEFT_IO["Description"].apply(extract_neft_category)
        df.update(NEFT_IO)

    def extract_net_neft_category(description):
        try:
            category_part = description.split("net-neft-")[1].split("-")[3]
            return category_part
        except IndexError:
            return "Suspense"

    NEFT = df[df["Description"].str.contains("net-neft-", na=False)]
    NEFT = NEFT[~NEFT["Category"].str.contains(
        "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
    if not NEFT.empty:
        NEFT["Category"] = NEFT["Description"].apply(extract_net_neft_category)
        df.update(NEFT)

    def extract_neft_name(description):
        try:
            name_part = description.split("neft-")[1].split("/")[0]
            return name_part
        except IndexError:
            return "Suspense"

    NEFT_Kar = df[df["Description"].str.contains("neft-", na=False) & df["Category"].str.contains("Suspense", na=False)]
    NEFT_Kar = NEFT_Kar[~NEFT_Kar["Category"].str.contains(
        "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
    if not NEFT_Kar.empty:
        NEFT_Kar["Category"] = NEFT_Kar["Description"].apply(extract_neft_name)
        df.update(NEFT_Kar)

    def extract_category_neft_sbi(description):
        if "totransfer-neft" in description:
            parts = description.split("-")
            name_part = parts[-1]
            return name_part
        else:
            try:
                parts = description.split("/")
                category_part = parts[2]
                if any(char.isdigit() for char in category_part):
                    return "Suspense"
                else:
                    return category_part
            except IndexError:
                return "Suspense"

    neft_SBI = df[df["Description"].str.contains("totransfer-neft", na=False)]
    neft_SBI = neft_SBI[~neft_SBI["Category"].str.contains(
        "Redemption, Dividend & Interest|Bank Interest Received|Salary Paid|Salary Received")]
    if not neft_SBI.empty:
        neft_SBI["Category"] = neft_SBI["Description"].apply(extract_category_neft_sbi)
        df.update(neft_SBI)

    def filter_emi_transactions(df):
        df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
        keywords = ["emi", "achidfcfblimited", "cholamandalaminvest", "lnpy", "ach/", "ach-", "achdr", "ecs", "achd",
                    "bajajfinance", "cms", "lamum",
                    "lcfm", "loanreco", "lptne", "nach", "magmafincorpltd", "toachdraditybirl", "toachdrambitfinv",
                    "toachdrclixcapita",
                    "toachdrdeutscheb", "toachdrdhaniloan", "toachdrfedbankfi", "toachdrfullerton", "toachdrindiabulls",
                    "toachdrindinfhouf",
                    "toachdrindusind", "toachdrlendingkar", "toachdrmagmafinco", "toachdrmahnimahin",
                    "toachdrmoneywisef", "toachdrneogrowth",
                    "toachdrtatacapita", "toachdrtpachmag", "toachdrtpachneo", "toachdrtpcapfrst", "toachdryesbankr",
                    "achracpc",
                    ]
        pattern = r"^(" + "|".join(keywords) + r")"
        emi_transactions = df[
            df["Description"].str.contains(pattern, case=False, regex=True) & (~df["Debit"].isnull()) & (
                        df["Debit"] > 0)]
        repeated_emi = emi_transactions[emi_transactions.duplicated(subset=["Debit"], keep=False)]
        repeated_emi = repeated_emi[~repeated_emi["Category"].str.contains("Investment", na=False)]
        filtered_emi_indices = repeated_emi[
            ~repeated_emi["Debit"].astype(str).str.endswith("000") & (repeated_emi["Debit"] > 1000)].index
        df.loc[filtered_emi_indices, "Category"] = "Probable EMI"

        return df

    filter_emi_transactions(df)

    def Bounce(df):
        keywords = ["return", "Bounce", "i/wchqreturn", "out-chqreturn"]
        pattern = r"\b(" + "|".join(keywords) + r")\b"
        df.loc[df["Description"].str.contains(pattern, regex=True) & (df["Debit"] > 0), "Category",] = "Bounce"

        return df

    Bounce(df)

    # Iterate through the rows of df2
    for _, keyword_row in df2.iterrows():
        mask = df["Description"].str.contains(keyword_row["Description"], case=False, na=False)
        if keyword_row["Debit / Credit"] == "Debit":
            mask = mask & (df["Debit"] > 0)  # check if Debit is greater than 0
        elif keyword_row["Debit / Credit"] == "Credit":
            mask = mask & (df["Credit"] > 0)  # check if Credit is greater than 0
        df.loc[mask, "Category"] = keyword_row["Category"]
    #####
    MPS = df[df["Description"].str.contains("mps/", na=False) & ~df["Description"].str.contains("imps/")]
    if not MPS.empty:
        for idx, row in MPS.iterrows():
            if row["Credit"] > 0:
                df.at[idx, "Category"] = "UPI-Cr"
            elif row["Debit"] > 0:
                df.at[idx, "Category"] = "UPI-Dr"

    Salary_credit = ((df["Description"].str.contains("imps|neft|rtgs", case=False, na=False)) & (
        df["Description"].str.contains("salary", case=False, na=False)) & (df["Credit"] > 0))
    Salary_debit = ((df["Description"].str.contains("imps|neft|rtgs", case=False, na=False)) & (
        df["Description"].str.contains("salary", case=False, na=False)) & (df["Debit"] > 0))
    df.loc[Salary_credit, "Category"] = "Salary Received"
    df.loc[Salary_debit, "Category"] = "Salary Paid"

    mask_withdrawal = (df["Description"].str.contains("eaw-|nwd-|atw-|tocash", case=False, na=False)) & (
                df["Debit"] > 0)
    df.loc[mask_withdrawal, "Category"] = "Cash Withdrawal"

    General_insurance = ["acko", "adityabirlahealth", "bajajallianz", "bhartiaxa", "carehealth", "cholamandalam",
                         "ecgc", "edelweiss",
                         "future generali", "godigit", "hdfcergo", "icicilombard", "iffcotokio", "kotakgeneral",
                         "liberty", "manipalcigna",
                         "maxbupahealth", "nationalinsurance", "pmsby", "rahejaqbe", "royalsundaram", "sbigeneral",
                         "shriram", "starhealth",
                         "tataaig", "thenewindiaassurance", "theoriental", "unitedindia", "universalsompo",
                         ]
    df.loc[df["Description"].apply(lambda x: any(keyword in x for keyword in General_insurance)) & (
                df["Debit"] > 0), "Category",] = "General insurance"

    online_shopping_keywords = ["amazon", "bigbasket", "ecom", "flipkart", "mamaearth", "myntra", "nykaa", "meesho", ]
    df.loc[df["Description"].apply(lambda x: any(
        keyword in x.lower() for keyword in online_shopping_keywords) and "amazonpay" not in x.lower()) & (
                       df["Debit"] > 0), "Category",] = "Online Shopping"

    INB = df[df["Description"].str.contains("inb/|inb-td/", na=False)]
    INB = INB[~INB["Description"].str.contains("gsttaxpayments", na=False)]
    INB = INB[~INB["Category"].str.contains("Salary Paid|Salary Received|GST Paid", na=False)]
    if not INB.empty:
        INB["Category"] = INB["Description"].apply(lambda x: (x.split("/")[2]
                                                              if "inb/" in x and len(x.split("/")) > 2
                                                              else x.split("/")[1] if len(x.split("/")) > 1 else x))

        df.update(INB)

    NEFT_INB = df[df["Description"].str.contains("inb/neft", na=False)]
    if not NEFT_INB.empty:
        NEFT_INB["Category"] = NEFT_INB["Description"].apply(
            lambda x: x.split("/")[3] if "inb/neft" in x and len(x.split("/")) > 3 else x)
        df.update(NEFT_INB)

    BIL_IMB_entries = df[df["Description"].str.contains("bil/imb/", na=False)]
    BIL_IMB_entries = BIL_IMB_entries[~BIL_IMB_entries["Category"].str.contains("Salary Paid,Salary Received")]
    if not BIL_IMB_entries.empty:
        def extract_name_from_bilimb(description):
            parts = description.split("/")
            if len(parts) >= 4:
                return parts[3]
            return None

        for idx, row in BIL_IMB_entries.iterrows():
            name = extract_name_from_bilimb(row["Description"])
            if name and not name.isdigit():
                df.at[idx, "Category"] = name
            else:
                df.at[idx, "Category"] = "Suspense"

    def extract_ecs_name(description):
        parts = description.split("/")
        if len(parts) > 1 and parts[1].strip():
            return parts[1].strip()
        else:
            return "Suspense"

    ECS = df[df["Description"].str.contains("ecs/", na=False)]
    ECS = ECS[~ECS["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not ECS.empty:
        ECS = ECS[~ECS["Category"].str.contains("Redemption|Dividend & Interest", na=False)]
        ECS = ECS[~ECS["Category"].str.contains("Probable EMI|Bank Charges", na=False)]
        ECS["Category"] = ECS["Description"].apply(extract_ecs_name)
        df.update(ECS)

    MPS = df[df["Description"].str.contains("MPS/", na=False)]
    if not ECS.empty:
        ECS = ECS[~ECS["Description"].str.contains("imps/")]

    def extract_imps_name(description):
        pattern = r"imps-\d+-(.*?)-"
        match = re.search(pattern, description)
        if match:
            extracted = match.group(1).strip()
            if extracted:
                return extracted
        return "Suspense"

    IMPS_HDFC = df[df["Description"].str.contains("imps", na=False)]
    IMPS_HDFC = IMPS_HDFC[~IMPS_HDFC["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not IMPS_HDFC.empty:
        IMPS_HDFC["Category"] = IMPS_HDFC["Description"].apply(extract_imps_name)
        df.update(IMPS_HDFC)

    def extract_imps_rib_name(description):
        parts = description.split("/")
        if len(parts) > 3 and parts[3].strip():
            return parts[3].strip()
        else:
            return "Suspense"

    imps_rib = df[df["Description"].str.contains("imps-rib|imps-inet|imps-cib", na=False)]
    imps_rib = imps_rib[~imps_rib["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not imps_rib.empty:
        imps_rib["Category"] = imps_rib["Description"].apply(extract_imps_rib_name)
        df.update(imps_rib)

    def extract_imps_mob_name(description):
        parts = description.split("/")
        if len(parts) > 3 and parts[3].strip():
            return parts[3].strip()
        else:
            return "Suspense"

    imps_mob = df[df["Description"].str.contains("imps-mob/", na=False)]
    imps_mob = imps_mob[~imps_mob["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not imps_mob.empty:
        imps_mob["Category"] = imps_mob["Description"].apply(extract_imps_mob_name)
        df.update(imps_mob)

    def extract_imps_hdfc_category(description):
        try:
            name_part = description.split("/")[2]
            if any(char.isdigit() for char in name_part):
                return "Suspense"
            return name_part
        except IndexError:
            return "Suspense"

    imps_idfc = df[df["Description"].str.contains("imps/", na=False)]
    imps_idfc = imps_idfc[~imps_idfc["Category"].str.contains("Salary Paid|Salary Received")]
    if not imps_idfc.empty:
        imps_idfc["Category"] = imps_idfc["Description"].apply(extract_imps_hdfc_category)
        df.update(imps_idfc)

    def extract_bulkposting_category(description):
        import re
        try:
            # Use regular expression to find the pattern between '_' and digits
            match = re.search(r'_(\D+?)(?=\d)', description)
            if match:
                return match.group(1)
            else:
                return "Suspense"
        except:
            return "Suspense"

    bulkposting_idfc = df[df["Description"].str.contains("bulkposting", na=False)]
    if not bulkposting_idfc.empty:
        bulkposting_idfc["Category"] = bulkposting_idfc["Description"].apply(extract_bulkposting_category)
        df.update(bulkposting_idfc)

    def extract_category_axis(x):
        try:
            category_part = x.split("/")[3]
            if any(char.isdigit() for char in category_part):
                return "Suspense"
            return category_part
        except IndexError:
            return "Suspense"

    imps_axis = df[df["Description"].str.contains("imps/p2a", na=False)]
    imps_axis = imps_axis[~imps_axis["Category"].str.contains("Salary Paid|Salary Received")]
    if not imps_axis.empty:
        imps_axis["Category"] = imps_axis["Description"].apply(extract_category_axis)
        df.update(imps_axis)

    def extract_category_axis(x):
        try:
            category_part = x.split("/")[3]
            category_cleaned = "".join(
                filter(lambda char: not char.isdigit(), category_part)
            )
            if not category_cleaned:
                return "Suspense"
            return category_cleaned
        except IndexError:
            return "Suspense"

    imps_axis = df[df["Description"].str.contains("mmt/imps", na=False)]
    imps_axis = imps_axis[~imps_axis["Category"].str.contains("Salary Paid|Salary Received")]
    if not imps_axis.empty:
        imps_axis["Category"] = imps_axis["Description"].apply(extract_category_axis)
        df.update(imps_axis)

    def extract_imps_fed_name(description):
        parts = description.split("/")
        if len(parts) > 3 and parts[3].strip():
            return parts[3].strip()
        else:
            return "Suspense"

    imps_fed = df[df["Description"].str.contains("ftimps", na=False)]
    imps_fed = imps_fed[~imps_fed["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not imps_fed.empty:
        imps_fed["Category"] = imps_fed["Description"].apply(extract_imps_fed_name)
        df.update(imps_fed)

    # imps_svc = df[df["Description"].str.contains("byimps", na=False)]
    # if not imps_svc.empty:
    #     imps_svc_name = imps_svc['Description'].apply(lambda x: x.split('-')[2])
    #     imps_svc['Category'] = imps_svc_name
    #     df.update(imps_svc)

    def safe_extract_category(description):
        try:
            return description.split("-")[2]
        except IndexError:
            return "Suspense"

    imps_svc = df[df["Description"].str.contains("byimps", na=False)]
    imps_svc = imps_svc[~imps_svc["Category"].str.contains("Salary Paid|Salary Received")]
    if not imps_svc.empty:
        imps_svc["Category"] = imps_svc["Description"].apply(safe_extract_category)
        df.update(imps_svc)

    # SBI
    def extract_category_imps_sbi(description):
        try:
            parts = description.split("/")
            category_part = parts[2]
            if any(char.isdigit() for char in category_part):
                return "Suspense"
            else:
                return category_part
        except IndexError:
            return "Suspense"

    imps_SBI = df[df["Description"].str.contains("totransfer-inbimps/p2a/", na=False)]
    imps_SBI = imps_SBI[~imps_SBI["Category"].str.contains("Salary Paid|Salary Received")]
    if not imps_SBI.empty:
        imps_SBI["Category"] = imps_SBI["Description"].apply(extract_category_imps_sbi)
        df.update(imps_SBI)

    def extract_mob_name(description):
        parts = description.split("/")
        if len(parts) > 2:
            if not parts[2].isdigit():
                candidate = parts[-2].strip() if len(parts) >= 2 else ""
                return candidate if candidate else "Suspense"
        return "Suspense"

    MOB = df[df["Description"].str.contains("mob/tpft/", na=False)]
    MOB = MOB[~MOB["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not MOB.empty:
        MOB["Category"] = MOB["Description"].apply(extract_mob_name)
        df.update(MOB)

    def extract_mob_1_name(description):
        parts = description.split("/")
        if len(parts) > 2:
            if not parts[2].isdigit():
                candidate = parts[-2].strip() if len(parts) >= 2 else ""
                return candidate if candidate else "Suspense"
        return "Suspense"

    MOB_1 = df[df["Description"].str.contains("mob/selfft/", na=False)]
    MOB_1 = MOB_1[~MOB_1["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not MOB_1.empty:
        MOB_1["Category"] = MOB_1["Description"].apply(extract_mob_1_name)
        df.update(MOB_1)

    def extract_brn_clg_name(description):
        try:
            part = description.split("to ")[-1]
            name = part.split("/")[0].strip()
            return name if name else "Suspense"
        except Exception:
            return "Suspense"

    BRN_clg = df[df["Description"].str.contains("brn-clg-chqpaidto", na=False)]
    if not BRN_clg.empty:
        BRN_clg["Category"] = BRN_clg["Description"].apply(extract_brn_clg_name)
        df.update(BRN_clg)

    def extract_brn_clg_name(description):
        try:
            # Split based on "brn-clg-chqpaidto" and then split on "/" to extract the category name.
            part = description.split("brn-clg-chqpaidto")[-1]
            name = part.split("/")[0].strip()
            return name if name else "Suspense"
        except Exception:
            return "Suspense"

    BRN_clg = df[df["Description"].str.contains("brn-clg-chqpaidto", na=False)]
    BRN_clg = BRN_clg[~BRN_clg["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not BRN_clg.empty:
        BRN_clg["Category"] = BRN_clg["Description"].apply(extract_brn_clg_name)
        df.update(BRN_clg)

    def extract_category_ben(x):
        try:
            category_part = x.split("/")[3]
            if any(char.isdigit() for char in category_part):
                return "Suspense"
            return category_part
        except IndexError:
            return "Suspense"

    imps_ben = df[df["Description"].str.contains("imps/ben/", na=False)]
    imps_ben = imps_ben[~imps_ben["Category"].str.contains("Salary Paid,Salary Received")]
    if not imps_ben.empty:
        imps_ben["Category"] = imps_ben["Description"].apply(extract_category_ben)
        df.update(imps_ben)

    def extract_sentimps_text_after_numeric(description):
        try:
            # Find the position of the "sentimps" prefix
            prefix_position = description.find("sentimps")
            if prefix_position != -1:
                # Extract the part after the prefix
                after_prefix = description[prefix_position + len("sentimps"):]
                # Find the first non-digit character after the numeric part
                first_non_digit_index = len(after_prefix)
                for i, char in enumerate(after_prefix):
                    if not char.isdigit():
                        first_non_digit_index = i
                        break
                # Extract the text part after the numeric sequence
                text_after_numeric = after_prefix[first_non_digit_index:]
                # Extract the part before the next '/'
                text_part = text_after_numeric.split("/")[0]
                return text_part if text_part else "Suspense"
            return "Suspense"
        except Exception as e:
            return "Suspense"

    imps_new = df[df["Description"].str.contains("sentimps", na=False)]
    imps_new = imps_new[~imps_new["Category"].str.contains("Salary Paid,Salary Received")]
    if not imps_new.empty:
        imps_new["Category"] = imps_new["Description"].apply(extract_sentimps_text_after_numeric)
        df.update(imps_new)

    Loan = df[df["Description"].str.contains("idfcfirstb|krazybees", na=False)]
    if not Loan.empty:
        df.loc[(df["Description"].str.contains("idfcfirstb|krazybees", regex=True)) & (
                    df["Credit"] > 0), "Category",] = "Loan"

    BRN = df[
        (df["Description"].str.contains("brn-flexi") | df["Description"].str.contains("/SBI Funds/STATE BAN//ATTN//")) &
        df["Credit"].notnull()]
    if not BRN.empty:
        df.loc[BRN.index, "Category"] = "Redemption of Investment"

    def extract_name(description):
        parts = description.split("/")
        if len(parts) > 2 and parts[2].strip():
            return parts[2].strip()
        else:
            return "Suspense"

    RTGS = df[df["Description"].str.contains("rtgs/", na=False)]
    RTGS = RTGS[~RTGS["Category"].str.contains("Salary Paid|Salary Received")]
    if not RTGS.empty:
        RTGS["Category"] = RTGS["Description"].apply(extract_name)
        df.update(RTGS)

    def extract_rtgs_category(description):
        try:
            return description.split("-")[2]
        except IndexError:
            return "Suspense"

    RTGS_HDFC_DR = df[
        df["Description"].str.contains("rtgsdr", na=False) | df["Description"].str.contains("rtgs-", na=False)]
    RTGS_HDFC_DR = RTGS_HDFC_DR[~RTGS_HDFC_DR["Category"].str.contains("Salary Paid,Salary Received")]
    if not RTGS_HDFC_DR.empty:
        RTGS_HDFC_DR["Category"] = RTGS_HDFC_DR["Description"].apply(extract_rtgs_category)
        df.update(RTGS_HDFC_DR)

    def extract_name_inrtgs(description):
        parts = description.split("/")
        if len(parts) > 1 and parts[1].strip():
            return parts[1].strip()
        else:
            return "Suspense"

    RTGS_DCB = df[df["Description"].str.contains("inrtgs", na=False)]
    RTGS_DCB = RTGS_DCB[~RTGS_DCB["Category"].str.contains("Salary Paid|Salary Received", na=False)]
    if not RTGS_DCB.empty:
        RTGS_DCB["Category"] = RTGS_DCB["Description"].apply(extract_name_inrtgs)
        df.update(RTGS_DCB)

    RTGS_DCB = df[df["Description"].str.contains("tortgs", na=False)]
    RTGS_DCB = RTGS_DCB[~RTGS_DCB["Category"].str.contains("Salary Paid,Salary Received")]
    if not RTGS_DCB.empty:
        def extract_category(description):
            try:
                return description.split("/")[1]
            except IndexError:
                return "Suspense"

        RTGS_name = RTGS_DCB["Description"].apply(extract_category)
        RTGS_DCB["Category"] = RTGS_name
        df.update(RTGS_DCB)

    toib_svc = df[df["Description"].str.contains("toib", na=False)]
    toib_svc = toib_svc[~toib_svc["Category"].str.contains("Salary Paid,Salary Received")]
    if not toib_svc.empty:
        def extract_category_svc(description):
            try:
                return description.split("-")[2]
            except IndexError:
                return "Suspense"

        toib_svc_names = toib_svc["Description"].apply(extract_category_svc)
        toib_svc["Category"] = toib_svc_names
        df.update(toib_svc)

    BYNEFTID_DCB = df[df["Description"].str.contains("byneftid", na=False)]
    BYNEFTID_DCB = BYNEFTID_DCB[~BYNEFTID_DCB["Category"].str.contains("Salary Paid,Salary Received")]
    if not BYNEFTID_DCB.empty:
        def extract_category(description):
            try:
                return description.split("from")[1].split("]")[0]
            except IndexError:
                return "Suspense"  # We return "Unknown" if the format is incorrect

        BYNEFTID_name = BYNEFTID_DCB["Description"].apply(extract_category)
        BYNEFTID_DCB["Category"] = BYNEFTID_name
        df.update(BYNEFTID_DCB)

    BYRTGSID_DCB = df[df["Description"].str.contains("byrtgsid", na=False)]
    BYRTGSID_DCB = BYRTGSID_DCB[~BYRTGSID_DCB["Category"].str.contains("Salary Paid,Salary Received")]
    if not BYRTGSID_DCB.empty:
        def extract_name_from_description(description):
            try:
                return "".join(
                    filter(str.isalpha, description.split("[")[-1].split("]")[0])
                )
            except IndexError:
                return "Suspense"  # We return "Unknown" if the format is incorrect

        BYRTGSID_name = BYRTGSID_DCB["Description"].apply(extract_name_from_description)
        BYRTGSID_DCB["Category"] = BYRTGSID_name
        df.update(BYRTGSID_DCB)

    BYCLG_DCB = df[df["Description"].str.contains("byclg:", na=False)]
    BYCLG_DCB = BYCLG_DCB[~BYCLG_DCB["Category"].str.contains("Salary Paid,Salary Received")]
    if not BYCLG_DCB.empty:
        def extract_name_from_description(description):
            try:
                name_part = description.split("byclg:")[1]
                name = "".join(filter(lambda x: not x.isdigit(), name_part))
                return name
            except IndexError:
                return "Suspense"  # We return "Unknown" if the format is incorrect

        BYCLG_name = BYCLG_DCB["Description"].apply(extract_name_from_description)
        BYCLG_DCB["Category"] = BYCLG_name
        df.update(BYCLG_DCB)

    CASHDEP = df[df["Description"].str.contains("cashdep|bytransfer|byclearing/|atr/|chequedeposit", na=False)]
    if not CASHDEP.empty:
        extracted_names = [
            re.search(r"CASHDEP(.*?)-", s).group(1)
            for s in CASHDEP["Description"]
            if re.search(r"CASHDEP(.*?)-", s)
        ]
        CASHDEP["Category"] = "Cash Deposits"
        df.update(CASHDEP)

    def extract_rtgs_name_jankalyan(description):
        if "rtgs" in description.lower():
            try:
                name_part = description.split("rtgs")[1]
                name_without_numbers = re.match(r"[a-zA-Z]+", name_part)
                if name_without_numbers:
                    return name_without_numbers.group(0)
                else:
                    return "Suspense"
            except IndexError:
                return "Suspense"
        # Removed the else condition

    RTGS_jankalyan = df[(df["Description"].str.contains("rtgs", case=False, na=False)) & (df["Category"] == "Suspense")]
    RTGS_jankalyan = RTGS_jankalyan[~RTGS_jankalyan["Category"].str.contains("Salary Paid,Salary Received")]
    if not RTGS_jankalyan.empty:
        RTGS_jankalyan["Category"] = RTGS_jankalyan["Description"].apply(extract_rtgs_name_jankalyan)
        df.update(RTGS_jankalyan)

    def extract_neft_jankalyan(description):
        if "neft" in description.lower():
            try:
                name_part = description.split("neft")[1]
                name_without_numbers = re.match(r"[a-zA-Z]+", name_part)
                if name_without_numbers:
                    return name_without_numbers.group(0)
                else:
                    return "Suspense"
            except IndexError:
                return "Suspense"

    NEFT = df[(df["Description"].str.contains("neft", case=False, na=False)) & (df["Category"] == "Suspense")]
    NEFT = NEFT[~NEFT["Category"].str.contains("Salary Paid,Salary Received")]
    if not NEFT.empty:
        NEFT["Category"] = NEFT["Description"].apply(extract_neft_jankalyan)
        df.update(NEFT)

    def extract_imps_category(description):
        if "imps[" in description.lower():
            try:
                # Splitting at 'imps[' and extracting the part after it
                parts = description.split("imps[")[1]
                # Further splitting by ']' to isolate the segments
                segments = parts.split("]")
                # The desired name is likely the second segment (after the first ']')
                if len(segments) > 1:
                    extracted_name = segments[1].strip("[").strip("]")
                    return extracted_name
                else:
                    return "Suspense"
            except IndexError:
                return "Suspense"

    IMPS = df[(df["Description"].str.contains("imps\[", case=False, na=False)) & (df["Category"] == "Suspense")]
    IMPS = IMPS[~IMPS["Category"].str.contains("Salary Paid,Salary Received")]
    if not IMPS.empty:
        IMPS["Category"] = IMPS["Description"].apply(extract_imps_category)
        df.update(IMPS)

    def extract_trtr_imps_category(description):
        try:
            parts = description.split("/")
            if len(parts) >= 4 and parts[0].lower() == "trtr" and parts[2].lower() == "imps":
                name_part = parts[3]  # Extract the part after "imps"
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part  # Return clean name
            else:
                return "Suspense"
        except IndexError:
            return "Suspense"

    imps_trtr = df[df["Description"].str.contains(r"trtr/\d+/imps/", case=False, na=False)].copy()
    imps_trtr = imps_trtr[~imps_trtr["Category"].str.contains("Salary Paid|Salary Received", case=False, na=False)]
    if not imps_trtr.empty:
        imps_trtr["Category"] = imps_trtr["Description"].apply(extract_trtr_imps_category)
        df.update(imps_trtr)  # Update original DataFrame

    def extract_imps_in_category(description):
        try:
            parts = description.split("/")
            if len(parts) >= 4 and parts[0].lower() == "imps-in":
                name_part = parts[3]  # Extract the part after the second number
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part  # Return the cleaned name
            else:
                return "Suspense"
        except IndexError:
            return "Suspense"

    imps_in_df = df[df["Description"].str.contains(r"imps-in/\d+/\d+/", case=False, na=False)].copy()
    imps_in_df = imps_in_df[
        ~imps_in_df["Category"].str.contains("Salary Paid|Salary Received", case=False, na=False)]
    if not imps_in_df.empty:
        imps_in_df["Category"] = imps_in_df["Description"].apply(extract_imps_in_category)
        df.update(
            imps_in_df)

    def extract_neft_in_category(description):
        try:
            parts = description.split("/")
            if len(parts) >= 3 and parts[0].startswith("neft_in:ioban"):
                name_part = parts[2]
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part
            else:
                return "Suspense"
        except IndexError:
            return "Suspense"

    neft_in_df = df[df["Description"].str.contains(r"neft_in:ioban\d+/\d+/", case=False, na=False)].copy()
    neft_in_df = neft_in_df[~neft_in_df["Category"].str.contains("Salary Paid|Salary Received", case=False, na=False)]
    if not neft_in_df.empty:
        neft_in_df["Category"] = neft_in_df["Description"].apply(extract_neft_in_category)
        df.update(neft_in_df)  # Update original DataFrame

    def extract_imps_io_category(description):
        try:
            parts = description.split("/")
            if len(parts) >= 4 and parts[0].lower() == "imps" and parts[1] in ["in", "out"]:
                name_part = parts[3]  # Extract the last part (actual name)
                if any(char.isdigit() for char in name_part):
                    return "Suspense"
                return name_part  # Return clean name
            else:
                return "Suspense"
        except IndexError:
            return "Suspense"

    imps_io_df = df[df["Description"].str.contains(r"imps/(in|out)/\d+/", case=False, na=False)].copy()
    imps_io_df = imps_io_df[~imps_io_df["Category"].str.contains("Salary Paid|Salary Received", case=False, na=False)]
    if not imps_io_df.empty:
        imps_io_df["Category"] = imps_io_df["Description"].apply(extract_imps_io_category)
        df.update(imps_io_df)  # Update original DataFrame

    PF = df[df["Description"].str.contains("providentfund", na=False)]
    if not PF.empty:
        df.loc[(df["Description"].str.contains("providentfund", regex=True)) & (
                    df["Credit"] > 0), "Category",] = "Provident Fund"

    Salary_credit = (
            (df["Description"].str.contains("imps", case=False, na=False)) & (
        df["Description"].str.contains("salary", case=False, na=False)) & (df["Credit"] > 0)
    )
    Salary_debit = (
            (df["Description"].str.contains("imps", case=False, na=False)) & (
        df["Description"].str.contains("salary", case=False, na=False)) & (df["Debit"] > 0)
    )
    df.loc[Salary_credit, "Category"] = "Salary Received"
    df.loc[Salary_debit, "Category"] = "Salary Paid"

    last_move = r"(imps|neft|rtgs|chqpaid|chqdep)"
    df.loc[(df["Description"].str.contains(last_move, regex=True)) & (df["Debit"] > 0) & (
                df["Category"] == "Suspense"), "Category",] = ""
    df.loc[(df["Description"].str.contains(last_move, regex=True)) & (df["Credit"] > 0) & (
                df["Category"] == "Suspense"), "Category",] = ""

    df["Voucher type"] = ""
    df.loc[df["Category"].isin(["Cash Withdrawal", "Cash Deposits"]), "Voucher type"] = "Contra"
    df.loc[(df["Debit"] > 0) & (~df["Category"].isin(["Cash Withdrawal", "Cash Deposits"])), "Voucher type"] = "Payment"
    df.loc[(df["Credit"] > 0) & (~df["Category"].isin(["Cash Withdrawal", "Cash Deposits"])), "Voucher type"] = "Receipt"
    df["Balance"] = x  # Manish
    return df


##SHEETS
def process_name_n_num_df(data):
    name_n_num_df = pd.DataFrame(data, columns=['Account Number', 'Account Name', 'Bank'])
    name_n_num_df = name_n_num_df.iloc[[0]]
    df_transposed = name_n_num_df.transpose()
    df_transposed.reset_index(inplace=True)
    df = pd.DataFrame(df_transposed)
    return df

def append_to_excel(file_path, new_data):
    sheet_name = "Sheet1"
    # Convert new data to a DataFrame
    new_df = pd.DataFrame(new_data)

    # Check if the file exists
    if os.path.exists(file_path):
        # Load existing Excel file
        with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            # Read existing sheet
            existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

            # Append new data
            updated_df = pd.concat([existing_df, new_df], ignore_index=True)

            # Write back to the same sheet
            updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        # Create a new Excel file with the given data
        new_df.to_excel(file_path, sheet_name=sheet_name, index=False)

    return file_path

def make_summary_great_again(df1, opening_closing_balance, df2):
    def generate_summary(table, value_column, summary_name):
        # Create pivot table
        summary = table.pivot_table(
            index="Category", columns="Month-Year", values=value_column, aggfunc="sum", fill_value=0
        )

        # Sort columns by date
        summary = summary.reindex(
            sorted(summary.columns, key=lambda x: pd.to_datetime(x, format='%b-%Y')), axis=1
        )

        # Add a total column
        summary["Total"] = summary.sum(axis=1)

        # Reset index and clean up
        summary = summary.reset_index()
        summary = summary[summary["Category"] != "X"]  # Remove rows with category "X"
        summary.rename(columns={"Category": summary_name}, inplace=True)
        return summary

    def get_missing_balances(opening_closing_balance, all_months):

        balances = {month: opening_closing_balance.get(month, ["", ""]) for month in all_months}

        # Fill missing balances
        previous_balance = None
        for month in all_months:
            opening, closing = balances[month]
            if opening == "":  # Fill missing opening balance
                balances[month][0] = previous_balance if previous_balance is not None else ""
            if closing == "":  # Fill missing closing balance
                # Find the next opening balance
                next_opening = None
                for next_month in all_months[all_months.index(month) + 1:]:
                    if balances[next_month][0] != "":
                        next_opening = balances[next_month][0]
                        break
                balances[month][1] = next_opening if next_opening is not None else balances[month][0]
            previous_balance = balances[month][1]

        # Sort results for output
        balances = dict(sorted(balances.items(), key=lambda x: pd.to_datetime(x[0], format='%b-%Y')))
        return balances

    def get_particulars(df1, balances):
        # Group by 'Month-Year' and calculate total credit and debit transactions
        summary = df1.groupby('Month-Year').agg(
            total_credit=('Credit', 'sum'),
            total_debit=('Debit', 'sum')
        ).reset_index()

        # Add opening and closing balances
        summary['opening_balance'] = summary['Month-Year'].map(lambda x: balances.get(x, [0, 0])[0])
        summary['closing_balance'] = summary['Month-Year'].map(lambda x: balances.get(x, [0, 0])[1])

        # Rearrange columns
        summary = summary[['Month-Year', 'opening_balance', 'closing_balance', 'total_credit', 'total_debit']]

        pivot_df = summary.set_index("Month-Year").T
        pivot_df["Total"] = pivot_df.sum(axis=1)
        pivot_df.index = [
            "Opening Balance",
            "Closing Balance",
            "Total Credit",
            "Total Debit",
        ]

        # Reset index for display
        date_columns = [col for col in pivot_df.columns if col != "Total"]

        # Sort date columns
        sorted_date_columns = sorted(date_columns, key=lambda x: pd.to_datetime(x, format="%b-%Y"))

        # Add 'Total' back to sorted columns
        sorted_columns = sorted_date_columns + ["Total"]

        # Reorder columns
        pivot_df = pivot_df[sorted_columns]

        # Sort columns by date
        pivot_df.reset_index(inplace=True)
        pivot_df.rename(columns={"index": "Particulars"}, inplace=True)

        return pivot_df

    category_add_rows = pd.DataFrame({
        "Value Date": [pd.to_datetime(df1["Value Date"], format='%d-%m-%Y').min()] * len(df2["Category"].unique()),
        "Description": ["null"] * len(df2["Category"].unique()),
        "Debit": [0] * len(df2["Category"].unique()),
        "Credit": [0] * len(df2["Category"].unique()),
        "Balance": [0] * len(df2["Category"].unique()),
        "Category": df2["Category"].unique(),
    })

    df1 = pd.concat([category_add_rows, df1]).reset_index(drop=True)

    # Adding 'Month-Year' column directly in the filtering and creation of the tables
    df1['Month-Year'] = pd.to_datetime(df1['Value Date'], format='%d-%m-%Y').dt.strftime('%b-%Y')

    all_month_years = pd.date_range(
        start=min([pd.to_datetime(key, format='%b-%Y') for key in opening_closing_balance.keys()]),
        end=max([pd.to_datetime(key, format='%b-%Y') for key in opening_closing_balance.keys()]),
        freq='MS'
    ).strftime('%b-%Y').tolist()


    # Creating a template row for each Month-Year in the complete range
    template_rows = pd.DataFrame({
        "Value Date": [
            pd.to_datetime("01-" + month_year, format='%d-%b-%Y').strftime('%d-%m-%Y')
            for month_year in all_month_years
        ],
        "Description": ["null"] * len(all_month_years),
        "Debit": [0] * len(all_month_years),
        "Credit": [0] * len(all_month_years),
        "Balance": [0] * len(all_month_years),
        "Category": ["X"] * len(all_month_years),
        "Month-Year": all_month_years,
    })

    df1 = pd.concat([template_rows, df1]).reset_index(drop=True)
    new_opening_closing_balance = get_missing_balances(opening_closing_balance, all_month_years)
    particulars_table = get_particulars(df1, new_opening_closing_balance)
    # df1["Value Date"] = pd.to_datetime(df1['Value Date'], format='%d-%m-%Y')

    income_table = df1[df1['Category'].isin(df2[df2['Particulars'] == 'Income']['Category'])]
    important_table = df1[df1['Category'].isin(df2[df2['Particulars'] == 'Important Expenses / Payments']['Category'])]
    other_table = df1[df1['Category'].isin(df2[df2['Particulars'] == 'Other Expenses / Payments']['Category'])]
    contra_credit_table = df1[df1['Category'].isin(df2[df2['Particulars'] == 'Contra Credit']['Category'])]
    contra_debit_table = df1[df1['Category'].isin(df2[df2['Particulars'] == 'Contra Debit']['Category'])]

    # Adding the template rows to each table and sorting them
    income_table = pd.concat([template_rows, income_table]).reset_index(drop=True)
    important_table = pd.concat([template_rows, important_table]).reset_index(drop=True)
    other_table = pd.concat([template_rows, other_table]).reset_index(drop=True)
    contra_credit_table = pd.concat([template_rows, contra_credit_table]).reset_index(drop=True)
    contra_debit_table = pd.concat([template_rows, contra_debit_table]).reset_index(drop=True)

    # Converting the "Value Date" column to datetime format in income_table, important_table, and other_table
    income_table['Value Date'] = pd.to_datetime(income_table['Value Date'], format='%d-%m-%Y')
    important_table['Value Date'] = pd.to_datetime(important_table['Value Date'], format='%d-%m-%Y')
    other_table['Value Date'] = pd.to_datetime(other_table['Value Date'], format='%d-%m-%Y')
    contra_credit_table['Value Date'] = pd.to_datetime(contra_credit_table['Value Date'], format='%d-%m-%Y')
    contra_debit_table['Value Date'] = pd.to_datetime(contra_debit_table['Value Date'], format='%d-%m-%Y')

    income_table = income_table.sort_values(by=['Value Date']).reset_index(drop=True)
    important_table = important_table.sort_values(by=['Value Date']).reset_index(drop=True)
    other_table = other_table.sort_values(by=['Value Date']).reset_index(drop=True)
    contra_credit_table = contra_credit_table.sort_values(by=['Value Date']).reset_index(drop=True)
    contra_debit_table = contra_debit_table.sort_values(by=['Value Date']).reset_index(drop=True)

    income_summary = generate_summary(income_table, "Credit", "Income / Receipts")
    important_summary = generate_summary(important_table, "Debit", "Important Expenses / Payments")
    other_summary = generate_summary(other_table, "Debit", "Other Expenses / Payments")
    contra_credit_summary = generate_summary(contra_credit_table, "Credit", "Contra Credit")
    contra_debit_summary = generate_summary(contra_debit_table, "Debit", "Contra Debit")

    missing_months_list = get_missing_months(opening_closing_balance, new_opening_closing_balance)

    def filter_non_defaulters(df1, df2):
        for row_number, row in df1.iterrows():
            if row['Total'] == 0:
                print("FOUND a")
                category = row.iloc[0]  # Extract Category from first column
                preference = df2.loc[df2['Category'] == category, 'Preferences'].values

                if len(preference) > 0 and 'non_default' in preference:
                    print("FOUND z")
                    df1.drop(index=row_number, inplace=True)

        df1.reset_index(drop=True, inplace=True)
        return df1

    income_summary = filter_non_defaulters(income_summary, df2)
    important_summary = filter_non_defaulters(important_summary, df2)
    other_summary = filter_non_defaulters(other_summary, df2)
    contra_credit_summary = filter_non_defaulters(contra_credit_summary, df2)
    contra_debit_summary = filter_non_defaulters(contra_debit_summary, df2)


    return particulars_table, income_summary, important_summary, other_summary, contra_credit_summary, contra_debit_summary, missing_months_list


def summary_sheet(idf, open_bal, close_bal, new_tran_df, new_categories = None):

    opening_closing_balance = {month: [open_bal[month], close_bal[month]] for month in open_bal}

    excel_file_path = os.path.join(BASE_DIR, "Final_Category.xlsx")
    user_created = os.path.join(BASE_DIR, "Customer_category.xlsx")
    # print("excel_file_path_bruh -",excel_file_path)
        # excel_file_path+user_created
    df2 = pd.read_excel(excel_file_path)
    user_created_df = pd.read_excel(user_created)
    
    df_new = pd.DataFrame()
    
    if new_categories:
        print("new_categories -",new_categories)
        df_new = pd.DataFrame(new_categories)
        append_to_excel(user_created, new_categories)

    # Append new data
    df2 = pd.concat([df2, df_new,user_created_df], ignore_index=True)

    sheet_1, sheet_2, sheet_3, sheet_4, sheet_5, sheet_6, missing_months_list = make_summary_great_again(new_tran_df, opening_closing_balance, df2)
    df_list = [sheet_1, sheet_2, sheet_3, sheet_4, sheet_5, sheet_6]

    return df_list, missing_months_list


def transaction_sheet( df):
    # print all cols of df
    if len(df["Bank"].unique()) > 1:
        tdf = df[
            [
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Entity",
                "Bank",
                "Voucher type"
            ]
        ]
    else:
        #1234_temp
        tdf = df[
            ["Value Date", "Description", "Debit", "Credit", "Balance", "Category","Entity","Bank","Voucher type"]
        ]
    return tdf

def total_investment( df):
    invest_df = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Investment":
            invest_df = invest_df._append(row, ignore_index=True)

    if invest_df.empty:
        invest_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return invest_df

def redemption_investment( df):
    red_df = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Redemption of Investment":
            red_df = red_df._append(row, ignore_index=True)

    if red_df.empty:
        red_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return red_df

def cash_withdraw( df):
    cashw_df = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Cash Withdrawal":
            cashw_df = cashw_df._append(row, ignore_index=True)

    if cashw_df.empty:
        cashw_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return cashw_df

def cash_depo( df):
    cashd_df = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Cash Deposits":
            cashd_df = cashd_df._append(row, ignore_index=True)

    if cashd_df.empty:
        cashd_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return cashd_df

def Bank_charges( df):
    Bank_df = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Bank Charges":
            Bank_df = Bank_df._append(row, ignore_index=True)

    if Bank_df.empty:
        Bank_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return Bank_df

def Entertainment( df):
    Entertainment = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Entertainment":
            Entertainment = Entertainment._append(row, ignore_index=True)

    if Entertainment.empty:
        Entertainment = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return Entertainment

def div_int( df):
    iii = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Redemption, Dividend & Interest":
            iii = iii._append(row, ignore_index=True)

    if iii.empty:
        iii = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return iii

def emi( df):
    em_i = pd.DataFrame()
    for index, row in df.iterrows():
        arow = row["Category"]
        if arow == "Probable EMI":
            em_i = em_i._append(row, ignore_index=True)

    if em_i.empty:
        em_i = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return em_i

def refund_reversal( df):
    refund = df[df["Category"].str.contains("Refund/Reversal", na=False)]

    if refund.empty:
        refund = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return refund


def creditor_list(df):
    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
    debit = df[
        (~df["Debit"].isnull()) & ((df["Credit"].isnull()) | (df["Credit"] == 0))
        ]
    patterns = [
        "toib-", "brn-clg-chq", "mmt/imps", "neftdr", "neft/mb/", "nft/", "mob/tpft", "nlcindialtd", "neft/mb/ax",
        "tortgs",
        "rtgsdr", "mob/tpft/", "imb/", "imps", "imps/p2a", "mob/selfft/", "inb/", "inb-", "chqpaid", "fundtrf", "iconn",
        "imps-cib", "imps-inet", "imps-rib", "imps-mob", "inft", "mbk/xfer", "neft", "payc", "r-utr", "vmt-icon",
        "chqpaid",
        "byclg", "rtgs", "neftn", "inb-", "neft-barb", "ecs/", "bulkposting"
    ]
    regex_pattern = "|".join(patterns)
    Creditor_list = debit[
        debit["Description"].str.contains(regex_pattern, case=False)
    ]

    def extract_name(description):
        match = re.match(r"^[a-zA-Z\s]+$", description)
        return match.group(0) if match else None

    name_transactions = df[
        (df["Category"] == "Creditor") & (df["Description"].apply(is_name))
        ]
    name_transactions["Category"] = name_transactions["Description"].apply(
        extract_name
    )
    Creditor_list = pd.concat([Creditor_list, name_transactions])

    # Additional code to exclude specified keywords
    exclude_keywords = ["ach/", "ach-", "achdr", "achd", "bajajfinance", "cms/", "lamum", "lcfm", "lnpy",
                        "loanreco", "lptne", "nach", "magmafincorpltd", "toachdraditybirl", "toachdrambitfinv",
                        "toachdrclixcapita", "toachdrdeutscheb", "toachdrdhaniloan", "toachdrfedbankfi",
                        "toachdrfullerton",
                        "toachdrindiabulls", "toachdrindinfhouf", "toachdrindusind", "toachdrlendingkar",
                        "toachdrmagmafinco",
                        "toachdrmahnimahin", "toachdrmoneywisef", "toachdrneogrowth", "toachdrtatacapita",
                        "toachdrtpachmag",
                        "toachdrtpachneo", "toachdrtpcapfrst", "toachdryesbankr", "gsttaxpayment","self-chqpaid"
                        ]
    exclude_pattern = "|".join(exclude_keywords)
    Creditor_list = Creditor_list[
        ~Creditor_list["Description"].str.contains(exclude_pattern, case=False)
    ]
    exclude_descriptions = ["billdesk", "gsttaxpayment", "atomstockbroker", "Probable Claim Settlement",
                            "Subscription / Entertainment"]
    for exclude in exclude_descriptions:
        Creditor_list = Creditor_list[
            ~Creditor_list["Description"].str.contains(exclude, case=False)
        ]
    exclude_categories = [
        "Payment Received",
        "Payment Made",
        "fastag",
        "Refund/Reversal",
        "Salary Paid",
        "Loan given",
        "Credit Card Payment",
        "Food Expense/Hotel",
        "Income Tax Paid",
        "Rent Paid",
        "Utility Bills",
        "Reimbursement",
        "Travelling Expense",
        "Bank Charges",
        "POS-Cr",
        "POS-Dr",
        "Payment Made",
        "Payment Received",
        "Cash Withdrawal",
        "Bonus Paid",
        "General insurance",
        "Investment",
        "Online Shopping",
        "Probable EMI",
        "TDS Deducted",
        "GST Paid",
    ]
    for exclude in exclude_categories:
        Creditor_list = Creditor_list[
            ~Creditor_list["Category"].str.contains(exclude, case=False)
        ]
    Creditor_list = Creditor_list.sort_values(by="Category")
    return Creditor_list

def debtor_list(df):
    df["Debit"] = pd.to_numeric(df["Debit"], errors="coerce")
    df["Credit"] = pd.to_numeric(df["Credit"], errors="coerce")
    credit = df[
        (~df["Credit"].isnull()) & ((df["Debit"].isnull()) | (df["Debit"] == 0))
        ]
    patterns = ["toib-", "neft", "mmt/imps", "neftcr", "imps", "tortgs", "rtgs", "rtgscr", "ecs/", "mob/tpft/", "imb/",
                "mob/selfft/",
                "inb/", "imps-mob", "nft/", "byclg", "inb-", "neft-", "googleindiadigital", "gsttaxpayment",
                "bulkposting", "chqdep"
                ]
    regex_pattern = "|".join(patterns)
    Debtor_list = credit[
        credit["Description"].str.contains(regex_pattern, case=False)
    ]

    exclude_categories = [
        "Redemption, Dividend & Interest",
        "Redemption of Investment",
        "Refund/Reversal",
        "Loan",
        "Provident Fund",
        "Payment Made",
        "Payment Received",
        "Bounce",
        "Reimbursement",
        "GST Paid",
        "Probable Claim Settlement",
        "Subscription / Entertainment"
    ]
    Debtor_list = Debtor_list[
        ~Debtor_list["Category"].str.contains(
            "|".join(exclude_categories), case=False
        )
    ]

    def extract_name(description):
        match = re.match(r"^[a-zA-Z\s]+$", description)
        return match.group(0) if match else None

    name_transactions = df[
        (df["Category"] == "Debtor") & (df["Description"].apply(is_name))
        ]
    name_transactions["Category"] = name_transactions["Description"].apply(
        extract_name
    )
    Debtor_list = pd.concat([Debtor_list, name_transactions])

    Debtor_list = Debtor_list.sort_values(by="Category")
    return Debtor_list

def Upi(df):
    categories_to_include = ["UPI-Cr", "UPI-Dr"]

    def apply_regex_to_empty_entities_axis(row):
        if row['Category'] in categories_to_include:
            if "upi/p2a" in row['Description'] or "upi/p2m" in row['Description']:
                match = re.search(r'upi/p2[am]/\d+/([^/]+)/', row['Description'])
                if match:
                    return match.group(1)
        return row['Entity']

    def apply_regex_to_categories_hdfc(row):
            if row['Category'] in categories_to_include:
                if "upi-" in row['Description']:
                    # For the simplest pattern (first type)
                    match1 = re.search(r'(?<=upi-)([a-zA-Z]+)', row['Description'])

                    # Comprehensive pattern that requires at least one letter in the name
                    match2 = re.search(r'upi-\d+-([a-zA-Z][a-zA-Z0-9._]*)[-@]', row['Description'])

                    if match2:
                        return match2.group(1)
                    elif match1:
                        return match1.group(1)

            return row['Entity']

    def apply_regex_to_empty_entities_sbi(row):
        if row['Category'] in categories_to_include:
            if "totransfer-upi" in row['Description'] or "bytransfer-upi" in row['Description']:
                match = re.search(r'(totransfer|bytransfer)-upi/[cd]r/\d+/([a-zA-Z]+)/', row['Description'])
                if match:
                    return match.group(2)
        return row['Entity']

    def apply_regex_to_empty_entities_kotak(row):
        if row['Category'] in categories_to_include:
            # Check if 'upi/' is in the description
            if "upi/" in row['Description']:
                # Match the name immediately after 'upi/'
                match = re.search(r'upi/([a-zA-Z.]+)', row['Description'])
                if match:
                    # Clean the name by removing special characters and numbers
                    name = re.sub(r'[^a-zA-Z]', '', match.group(1))  # Keep only alphabetic characters
                    return name if name else "Suspense"  # Return 'Suspense' if the name is empty
        return row['Entity']

    def apply_regex_to_empty_entities_rbl(row):
        if row['Category'] in categories_to_include:
            if "upi/" in row['Description']:
                match = re.search(r'upi/\d+/\w+/([a-zA-Z0-9@.]+)', row['Description'])
                if match:
                    # Extract the raw name
                    raw_name = match.group(1)
                    # Remove numeric and '@' characters
                    cleaned_name = re.sub(r'[0-9@]', '', raw_name)
                    return cleaned_name
        return row['Entity']

    def apply_regex_to_empty_entities_idfc(row):
        if row['Category'] in categories_to_include:
            if "upi/mob" in row['Description']:
                match = re.search(r'upi/mob/\d+/([\w]+)', row['Description'])
                if match:
                    raw_name = match.group(1)
                    cleaned_name = re.sub(r'[0-9@]', '', raw_name)
                    print(f"Extracted Name: {cleaned_name}")
                    return cleaned_name
        return row['Entity']

    def apply_regex_to_empty_entities_vasai(row):
        if row['Category'] in categories_to_include:
            match = re.search(r'upi/(cr|dr)/\d+/([\w]+)/', row['Description'])
            if match:
                extracted_name = match.group(2)
                return extracted_name
        return row['Entity']

        # Step 1: Apply regex logic first

    def extract_name_upiab(row):
        if row['Category'] in categories_to_include:
            match = re.search(r'upiab/\d+/cr/([\w]+)/', row['Description'])
            if match:
                extracted_name = match.group(1)  # Use group(1) for the first capturing group
                return extracted_name.capitalize()  # Capitalize the name for consistency
        return row['Entity']

    def extract_name_mpay(row):
        if row['Category'] in categories_to_include and row['Description'].startswith("mpay/upi/"):
            match = re.search(r'mpay/upi/.+?/\w+/([\w]+)+@', row['Description'])
            if match:
                extracted_name = match.group(1)
                cleaned_name = re.sub(r'[0-9@]', '', extracted_name)
                return cleaned_name
        return row['Entity']

    def apply_regex_to_categories_jsbl(row):
        if row['Category'] in categories_to_include:
            pattern = r"upi/(?:cr|dr)/\d+/([^/]+)"
            match = re.search(pattern, row['Description'])
            if match:
                return match.group(1)  # group(1) is the name
        return row['Entity']

    def apply_regex_to_categories_dcb(row):
        if row['Category'] in categories_to_include:
            # Modified pattern to match "upi:pay:" or "upi:rec:" format
            pattern = r"upi:(?:pay|rec):\d+/([^/]+)"
            match = re.search(pattern, row['Description'])
            if match:
                return match.group(1)  # Extract the name
        return row['Entity']

    def apply_regex_to_categories_idfc(row):
        if row['Category'] in categories_to_include:
            pattern = r"upi/mob/\d+/([^/]+)"  # Added 'mob' to the pattern
            match = re.search(pattern, row['Description'])
            if match:
                return match.group(1)  # Returns the captured name
        return row['Entity']

    def apply_regex_to_categories_uco(row):
        if row['Category'] in categories_to_include:
            new_pattern = r"(?:mpay/)?upi/trtr/\d+/[^/]+/([^/.]+)"
            new_match = re.search(new_pattern, row['Description'])
            if new_match:
                return new_match.group(1)  # Returns 'dream11'
        return row['Entity']

    def apply_regex_to_categories_nkgsb(row):
        if row['Category'] in categories_to_include:
            pattern1 = r"upi/(?:credit|debit)/\d+/([^/]+)"
            pattern2 = r"upi/(?:credit|debit)/([^/]+/\d+/)"
            match1 = re.search(pattern1, row['Description'])
            if match1:
                return match1.group(1)  # Returns the captured name
            match2 = re.search(pattern2, row['Description'])
            if match2:
                return match2.group(1).split('/')[0]  # Extracts only the name part
        return row['Entity']

    def apply_regex_to_categories_surat(row):
        if row['Category'] in categories_to_include:
            pattern2 = r"upi/(?:credit|debit)/([^/]+/\d+/)"
            match1 = re.search(pattern2, row['Description'])
            if match1:
                return match1.group(1)
        return row['Entity']


    df['Entity'] = df.apply(apply_regex_to_categories_uco, axis=1)
    df['Entity'] = df.apply(apply_regex_to_categories_nkgsb, axis=1)
    df['Entity'] = df.apply(apply_regex_to_categories_surat, axis=1)
    df['Entity'] = df.apply(apply_regex_to_categories_idfc, axis=1)
    df['Entity'] = df.apply(apply_regex_to_categories_jsbl, axis=1)   
    df['Entity'] = df.apply(apply_regex_to_categories_hdfc, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_sbi, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_kotak, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_rbl, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_idfc, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_vasai, axis=1)
    df['Entity'] = df.apply(extract_name_upiab, axis=1)
    df['Entity'] = df.apply(extract_name_mpay, axis=1)
    df['Entity'] = df.apply(apply_regex_to_categories_dcb, axis=1)
    df['Entity'] = df.apply(apply_regex_to_empty_entities_axis, axis=1)

    # print(df)
    return df

# def is_name( description):
#     return bool(re.match(r"^[a-zA-Z\s]+$", description))

def is_name(description):
    # Adjusted to strip whitespace and check for word characters
    description = description.strip()
    return bool(re.match(r"^[a-zA-Z\s]+$", description))

# def categorize_name_transactions( df):
#     # Apply 'NameTransaction' category for descriptions that appear to be names
#     df.loc[
#         (
#             df["Category"].isin(["Suspense", "Payment Made", "Payment Received"])
#         )  # Check if Category is either 'Suspense' or 'Payment Made'
#         & df["Description"].apply(
#             is_name
#         )  # Descriptions that appear to be names
#         & (
#             df["Category"] != "NameTransaction"
#         ),  # Ensure it's not already categorized as 'NameTransaction'
#         "Category",
#     ] = "NameTransaction"
#     df.loc[
#         (
#             df["Category"] == "NameTransaction"
#         )  # Consider only 'NameTransaction' category
#         & df["Credit"].notna(),  # Credit transactions
#         "Category",
#     ] = "Debtor"
#     df.loc[
#         (
#             df["Category"] == "NameTransaction"
#         )  # Consider only 'NameTransaction' category
#         & df["Debit"].notna(),  # Debit transactions
#         "Category",
#     ] = "Creditor"
#
#     return df

def categorize_name_transactions(df):
    # Clean the Description field
    df["Description"] = df["Description"].str.replace(r"[^a-zA-Z\s]", "", regex=True).str.strip()

    # Ensure Debit and Credit are numeric
    df["Debit"] = pd.to_numeric(df["Debit"], errors='coerce')
    df["Credit"] = pd.to_numeric(df["Credit"], errors='coerce')

    # Apply 'NameTransaction' category for descriptions that appear to be names
    df.loc[
        df["Category"].isin(["Suspense", "Payment Made", "Payment Received"])
        & df["Description"].apply(is_name)
        & (df["Category"] != "NameTransaction"),
        "Category",
    ] = "NameTransaction"

    # Further categorize as 'Debtor' or 'Creditor'
    df.loc[
        (df["Category"] == "NameTransaction") & df["Credit"].notna(),
        "Category",
    ] = "Debtor"
    df.loc[
        (df["Category"] == "NameTransaction") & df["Debit"].notna(),
        "Category",
    ] = "Creditor"

    return df


def another_method(df):
    # Categorize transactions as Creditor and Debtor
    Creditor_list = creditor_list(df)
    Debtor_list = debtor_list(df)
    NEW_DF = df.copy()
    NEW_DF["Entity"] = ''
    Creditor_list["Entity"] = Creditor_list["Category"]
    Debtor_list["Entity"] = Debtor_list["Category"]
    Creditor_list["Category"] = "Creditor"
    Debtor_list["Category"] = "Debtor"
    NEW_DF.update(Creditor_list)
    NEW_DF.update(Debtor_list)
    # NEW_DF.loc[(NEW_DF["Description"].str.contains("UPI", case=False)) & (NEW_DF["Debit"] > 0),"Category"] = "UPI-Dr"
    # NEW_DF.loc[(NEW_DF["Description"].str.contains("UPI", case=False)) & (NEW_DF["Credit"] > 0),"Category",] = "UPI-Cr"
    
    return NEW_DF


def suspense_credit( df):
    c_df = pd.DataFrame()
    for index, row in df.iterrows():
        credit_amount = pd.to_numeric(row["Credit"], errors="coerce")
        arow = row["Category"]
        if arow == "Suspense" and credit_amount > 0:
            c_df = c_df._append(row, ignore_index=True)
    if c_df.empty:
        c_df = pd.DataFrame(columns=["Value Date", "Description", "Credit"])
    else:
        c_df = c_df[["Value Date", "Description", "Credit"]]
    return c_df

def suspense_debit( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
        arow = row["Category"]
        if arow == "Suspense" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(columns=["Value Date", "Description", "Debit"])
    else:
        d_df = d_df[["Value Date", "Description", "Debit"]]
    return d_df

def payment( df):
    df_debit = df[df["Debit"].notnull()]
    iii = pd.DataFrame()
    for index, row in df_debit.iterrows():
        new_row = {
            "Date": row["Value Date"],
            "Effective Date": "",
            "Bill Ref": "",
            "Dr Ledger": row["Category"],
            "Cr Ledger": "",
            "Amount": row["Debit"],
            "Narration": row[
                "Description"
            ],  # Assuming you want to keep Narration as the description
        }
        iii = iii._append(new_row, ignore_index=True)
    if iii.empty:
        iii = pd.DataFrame(
            columns=[
                "Date",
                "Effective Date",
                "Bill Ref",
                "Dr Ledger",
                "Cr Ledger",
                "Amount",
                "Narration",
            ]
        )
    return iii

def receipt( df):
    df_credit = df[df["Credit"].notnull()]
    iii = pd.DataFrame()
    for index, row in df_credit.iterrows():
        new_row = {
            "Date": row["Value Date"],
            "Effective Date": "",
            "Cr Ledger": row["Category"],
            "Dr Ledger": "",
            "Amount": row["Credit"],
            "Narration": row[
                "Description"
            ],  # Assuming you want to keep Narration as the description
        }
        iii = iii._append(new_row, ignore_index=True)
    if iii.empty:
        iii = pd.DataFrame(
            columns=[
                "Date",
                "Effective Date",
                "Cr Ledger",
                "Dr Ledger",
                "Amount",
                "Narration",
            ]
        )

    return iii

def BOUNCE( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
        arow = row["Category"]
        if arow == "Bounce" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return d_df

def BankwiseEli( data, eod):

    error_df = pd.DataFrame()
    processed_data = calculate_fixed_day_average(eod)
    # Check for the existence of 'Avg_Last_6_Months' column
    if "Avg_Last_6_Months" not in processed_data.columns:
        print("'Avg_Last_6_Months' column not found in processed_data.")
        return error_df  # Or handle the absence of this column as needed

    # Conditional handling for 'Avg_Last_12_Months'
    if "Avg_Last_12_Months" in processed_data.columns:
        selected_columns = processed_data[
            ["Day", "Avg_Last_12_Months", "Avg_Last_6_Months"]
        ]
    else:
        print(
            "'Avg_Last_12_Months' column not found, proceeding with 'Avg_Last_6_Months' only."
        )
        selected_columns = processed_data[["Day", "Avg_Last_6_Months"]]

    # For Axis
    if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 1:
        if not pd.isna(data["Avg_Last_6_Months"].iloc[1]):
            avg_divided_by_1_5 = data["Avg_Last_6_Months"].iloc[1] / 1.5
        else:
            avg_divided_by_1_5 = np.nan
    else:
        print(
            "'Avg_Last_6_Months' column does not exist or does not have enough data."
        )
        avg_divided_by_1_5 = np.nan

    # For 'Avg_Last_12_Months' at index 0
    if "Avg_Last_12_Months" in data.columns:
        if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
            avg_divided_by_2_idfc = data["Avg_Last_12_Months"].iloc[0] / 2
        else:
            avg_divided_by_2_idfc = np.nan
    else:
        print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
        avg_divided_by_2_idfc = np.nan

    # For 'Avg_Last_6_Months' at index 2
    if "Avg_Last_6_Months" in data.columns and len(data["Avg_Last_6_Months"]) > 2:
        if not pd.isna(data["Avg_Last_6_Months"].iloc[2]):
            avg_divided_by_2_indus = data["Avg_Last_6_Months"].iloc[2] / 1.5
        else:
            avg_divided_by_2_indus = np.nan
    else:
        print(
            "'Avg_Last_6_Months' column does not exist or does not have enough data."
        )
        avg_divided_by_2_indus = np.nan

    # For 'Avg_Last_12_Months' at index 0 again
    if "Avg_Last_12_Months" in data.columns:
        if not pd.isna(data["Avg_Last_12_Months"].iloc[0]):
            avg_divided_by_2_L_T = data["Avg_Last_12_Months"].iloc[0] / 2
        else:
            avg_divided_by_2_L_T = np.nan
    else:
        print("'Avg_Last_12_Months' column does not exist in the DataFrame.")
        avg_divided_by_2_L_T = np.nan

    annual_interest_rate = 0.0870
    periods = 20 * 12
    principal = 100000
    payment_value = pmt(principal, annual_interest_rate, periods)
    payment_for_lap = pmt_lap()
    payment_for_bl = pmt_bl()

    # Calculating Loan value for axis
    axis_home_loan_value = None
    if payment_value != 0:
        axis_home_loan_value = avg_divided_by_1_5 / payment_value
        axis_home_loan_value = axis_home_loan_value * 100000
        axis_home_loan_value = round(axis_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    axis_LAP_value = None
    if payment_for_lap != 0:
        axis_LAP_value = avg_divided_by_1_5 / payment_for_lap
        axis_LAP_value = axis_LAP_value * 100000
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    axis_bl_value = None
    if payment_for_bl != 0:
        axis_bl_value = avg_divided_by_1_5 / payment_for_bl
        axis_bl_value = axis_bl_value / payment_for_lap
        axis_bl_value = axis_bl_value * 100000
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    # Calculating loan value for Idfc
    Idfc_home_loan_value = None
    if payment_value != 0:
        Idfc_home_loan_value = avg_divided_by_2_idfc / payment_value
        Idfc_home_loan_value = Idfc_home_loan_value * 100000
        Idfc_home_loan_value = round(Idfc_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    Idfc_LAP_value = None
    if payment_for_lap != 0:
        Idfc_LAP_value = avg_divided_by_2_idfc / payment_for_lap
        Idfc_LAP_value = Idfc_LAP_value * 100000
        Idfc_LAP_value = round(Idfc_LAP_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    Idfc_bl_value = None
    if payment_for_bl != 0:
        Idfc_bl_value = avg_divided_by_2_idfc / payment_for_bl
        Idfc_bl_value = Idfc_bl_value * 100000
        Idfc_bl_value = round(Idfc_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    # Calculating loan value for Indus
    indus_home_loan_value = None
    if payment_value != 0:
        indus_home_loan_value = avg_divided_by_2_indus / payment_value
        indus_home_loan_value = indus_home_loan_value * 100000
        indus_home_loan_value = round(indus_home_loan_value, 2)

    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    indus_LAP_value = None
    if payment_for_lap != 0:
        indus_LAP_value = avg_divided_by_2_indus / payment_for_lap
        indus_LAP_value = indus_LAP_value * 100000
        indus_LAP_value = round(indus_LAP_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    indus_bl_value = None
    if payment_for_bl != 0:
        indus_bl_value = avg_divided_by_2_indus / payment_for_bl
        indus_bl_value = indus_bl_value * 100000
        indus_bl_value = round(indus_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    L_T_home_loan_value = None
    if payment_value != 0:
        L_T_home_loan_value = avg_divided_by_2_L_T / payment_value
        L_T_home_loan_value = L_T_home_loan_value * 100000
        L_T_home_loan_value = round(L_T_home_loan_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")
    L_T_LAP_value = None
    if payment_for_lap != 0:
        L_T_LAP_value = avg_divided_by_2_L_T / payment_for_lap
        L_T_LAP_value = L_T_LAP_value * 100000
        L_T_LAP_value = round(L_T_LAP_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    L_T_bl_value = None
    if payment_for_lap != 0:
        L_T_bl_value = avg_divided_by_2_L_T / payment_for_bl
        L_T_bl_value = L_T_bl_value * 100000
        L_T_bl_value = round(L_T_bl_value, 2)
    else:
        print("PMT calculation resulted in zero, cannot divide by zero.")

    # Adding new columns in the specific order, with differentiating spaces for the repeated names
    selected_columns["Bank / NBFC_1"] = " "  # Added _1 for differentiation
    selected_columns["Home Loan Eligibility"] = " "
    selected_columns["Bank / NBFC_2"] = " "  # Added _2 for differentiation
    selected_columns["LAP Loan Eligibility"] = " "
    selected_columns["Bank / NBFC_3"] = " "  # Added _3 for differentiation
    selected_columns["Loan Eligibility Business loan"] = " "

    # Axis and L&T
    selected_columns.at[1, "Bank / NBFC_1"] = "AXIS , L&t"
    selected_columns.at[1, "Home Loan Eligibility"] = (
        f"{axis_home_loan_value} , {L_T_home_loan_value}"
    )

    selected_columns.at[1, "Bank / NBFC_2"] = "AXIS"
    selected_columns.at[1, "LAP Loan Eligibility"] = axis_LAP_value

    selected_columns.at[1, "Bank / NBFC_3"] = "AXIS"
    selected_columns.at[1, "Loan Eligibility Business loan"] = axis_bl_value

    # IDFC
    selected_columns.at[0, "Bank / NBFC_1"] = "IDFC"
    selected_columns.at[0, "Home Loan Eligibility"] = Idfc_home_loan_value

    selected_columns.at[0, "Bank / NBFC_2"] = "IDFC , L&T"
    selected_columns.at[0, "LAP Loan Eligibility"] = (
        f"{Idfc_LAP_value} , {L_T_LAP_value}"
    )

    selected_columns.at[0, "Bank / NBFC_3"] = "IDFC , L&T"
    selected_columns.at[0, "Loan Eligibility Business loan"] = (
        f"{Idfc_bl_value} , {L_T_bl_value}"
    )

    # Indus
    selected_columns.at[2, "Bank / NBFC_1"] = "INDUS"
    selected_columns.at[2, "Home Loan Eligibility"] = indus_home_loan_value

    selected_columns.at[2, "Bank / NBFC_2"] = "INDUS"
    selected_columns.at[2, "LAP Loan Eligibility"] = indus_LAP_value

    selected_columns.at[2, "Bank / NBFC_2"] = "INDUS"
    selected_columns.at[2, "LAP Loan Eligibility"] = indus_bl_value

    return selected_columns

def Bl_eligibility_bankwise( trans, data, eod):
    payment_for_bl = pmt_bl()  # Factor used in loan eligibility calculation
    # print(payment_for_bl)
    # print("#" * 100)

    # Convert "Value Date" to datetime and add "YearMonth" column for monthly grouping
    trans['Value Date'] = pd.to_datetime(trans['Value Date'], format='%d-%m-%Y', errors='coerce')
    trans['YearMonth'] = trans['Value Date'].dt.to_period('M')

    # Filter transactions for "Probable EMI" to calculate the monthly obligation amount
    probable_emi = trans[trans['Category'] == 'Probable EMI']
    obligation = 0
    if not probable_emi.empty:
        monthly_emi = probable_emi.groupby('YearMonth')['Debit'].sum().reset_index()
        monthly_emi.columns = ['Month', 'Total EMI Amount']
        obligation = monthly_emi.iloc[-1]['Total EMI Amount'] if not monthly_emi.empty else 0

    # Initialize a list to store results for each bank
    bankwise_results = []

    # Calculate ABB using eod data
    processed_data = calculate_fixed_day_average(eod)
    # print(processed_data)

    # Helper Functions for ABB Extraction
    def filter_12_months_average(data):
        """Filter for the 12-month average data, if available."""
        if 'Avg_Last_12_Months' not in data.columns:
            print("Column 'Avg_Last_12_Months' is missing. Returning an empty DataFrame.")
            return pd.DataFrame()  # Return an empty DataFrame
        return data[['Day', 'Avg_Last_12_Months']]

    def extract_abb(data, day_key):
        """Extract ABB based on a specific day key."""
        if data.empty:
            print("DataFrame is empty. Cannot extract ABB.")
            return None
        if day_key not in data['Day'].values:
            print(f"'{day_key}' not found in the 'Day' column.")
            return None
        return data.loc[data['Day'] == day_key, 'Avg_Last_12_Months'].values[0]

    # Filter and Extract Default ABB for general use
    filtered_data = filter_12_months_average(processed_data)
    default_ABB = extract_abb(filtered_data, 'Daily_Avg')  # Default ABB for most banks

    # Define Bank Entries with specific AVERAGE BANKING and any custom ABB keys
    bank_entries = [
        {'Bank': 'CHOLA', 'DATE': '1 TO 30', 'AVERAGE BANKING': 200000},
        {'Bank': 'L & T', 'DATE': '1 TO 30', 'AVERAGE BANKING': 150000},
        {'Bank': 'CLIX', 'DATE': '1 TO 30', 'AVERAGE BANKING': default_ABB * 2 if default_ABB else 0},
        {'Bank': 'HERO', 'DATE': '1,5,15,20,25', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_1_5_15_20_25'},
        {'Bank': 'ABFL', 'DATE': '5,10,15,20,25', 'AVERAGE BANKING': 150000,'custom_day_key': 'Avg_Days_5_10_15_20_25'},
        {'Bank':'INDUSIND','DATE':'4,5,7,10,15,25','AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_4_5_7_10_15_25'},
        {'Bank': 'SHRIRAM', 'DATE': '1 To 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'AXIS', 'DATE': '1 To 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'ICICI', 'DATE': '5,15,25', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_5_15_25'},
        {'Bank': 'AXIS FINANCE', 'DATE': '1 To 30', 'AVERAGE BANKING': 200000},
        {'Bank': 'AMBIT', 'DATE': '5,10,15,20,25,30', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_5_10_15_20_25_30'},
        {'Bank': 'EDELWIESS', 'DATE': '5,10,15,20,26', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_5_10_15_20_26'},
        {'Bank': 'UNITY', 'DATE': '1 TO 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'LOAN FRAME', 'DATE': '1 TO 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'NEOGROWTH', 'DATE': '1 TO 30', 'AVERAGE BANKING': 200000},
        {'Bank': 'PROTIUM', 'DATE': '1 TO 30', 'AVERAGE BANKING': 200000},
        {'Bank': 'TATA', 'DATE': '1 TO 30', 'AVERAGE BANKING': 300000},
        {'Bank': 'KOTAK', 'DATE': '1 TO 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'POONAWALA', 'DATE': '1 TO 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'CREDIT SAISON', 'DATE': '1,5,10,18,25', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_1_5_18_25'},
        {'Bank': 'YES BANK', 'DATE': '5,15,25', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_5_15_25'},
        {'Bank': 'DB', 'DATE': '1,5,10,15,20,25', 'AVERAGE BANKING': 360000, 'custom_day_key': 'Avg_Days_1_5_10_15_20_25'},
        {'Bank': 'SMC', 'DATE': '1 TO 30', 'AVERAGE BANKING': 300000},
        {'Bank': 'MAS', 'DATE': '1 TO 30', 'AVERAGE BANKING': 150000},
        {'Bank': 'BAJAJ', 'DATE': '2,10,20,30', 'AVERAGE BANKING': 100000,'custom_day_key': 'Avg_Days_1_5_10_15_20_25'},
        {'Bank': 'IDFC', 'DATE': '1 TO 30', 'AVERAGE BANKING': 150000},
        {'Bank': 'SMFG - fulltron', 'DATE': '1 TO 30', 'AVERAGE BANKING': 150000},
        {'Bank': 'HDFC', 'DATE': '1,5,10,15,20,25', 'AVERAGE BANKING': 125000,'custom_day_key': 'Avg_Days_1_5_10_15_20_25'},
        {'Bank': 'FED', 'DATE': '1,5,10,15,20,25', 'AVERAGE BANKING': 300000,'custom_day_key': 'Avg_Days_1_5_10_15_20_25'},
        {'Bank': 'SCB', 'DATE': '1 TO 30', 'AVERAGE BANKING': 100000},
        {'Bank': 'GODREJ', 'DATE': '5,10,15,20,25,30', 'AVERAGE BANKING': 300000,'custom_day_key': 'Avg_Days_5_10_15_20_25_30'},
        {'Bank': 'UGRO', 'DATE': '1 TO 30', 'AVERAGE BANKING': 7000000},

    ]

    # Process each bank entry for eligibility
    for entry in bank_entries:
        bank_name = entry['Bank']
        date_range = entry['DATE']
        average_banking = entry['AVERAGE BANKING']

        # Use a custom ABB if defined for this bank; otherwise, use the default ABB
        day_key = entry.get('custom_day_key', 'Daily_Avg')
        ABB = extract_abb(filtered_data, day_key)

        # Calculate metrics if ABB is available
        if ABB is not None:
            NetABB = ABB - obligation
            Emi_proposed = NetABB * 0.25
            Eligible_Loan_Amount = Emi_proposed / payment_for_bl
        else:
            NetABB = Emi_proposed = Eligible_Loan_Amount = 0

        # Determine Eligibility Status based on Net ABB and AVERAGE BANKING
        eligibility_status = "Eligible" if NetABB >= average_banking else "Not Eligible"

        # Append the results for each bank
        bankwise_results.append({
            'Bank': bank_name,
            'DATE': date_range,
            'ABB': ABB,
            'Obligation': obligation,
            'Net ABB': NetABB,
            'AVERAGE BANKING': average_banking,
            'Emi of the proposed loan': Emi_proposed,
            'Eligible Loan Amount': Eligible_Loan_Amount,
            'Eligibility Status': eligibility_status
        })

    # Convert the list of results to a DataFrame for easier handling
    bankwise_df = pd.DataFrame(bankwise_results)
    # print(bankwise_df)

    return bankwise_df

def Pos_cr( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
        arow = row["Category"]
        if arow == "POS-Cr" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return d_df

def Pos_dr( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Debit"], errors="coerce")
        arow = row["Category"]
        if arow == "POS-Dr" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return d_df

def UPI_cr( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
        arow = row["Category"]
        if arow == "UPI-Cr" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return d_df

def UPI_dr( df):
    d_df = pd.DataFrame()
    for index, row in df.iterrows():
        debit_amount = pd.to_numeric(row["Credit"], errors="coerce")
        arow = row["Category"]
        if arow == "UPI-Dr" and debit_amount > 0:
            d_df = d_df._append(row, ignore_index=True)
    if d_df.empty:
        d_df = pd.DataFrame(
            columns=[
                "Value Date",
                "Description",
                "Debit",
                "Credit",
                "Balance",
                "Category",
                "Bank",
            ]
        )
    return d_df

##EXCEL SHEETS MANIPULATION
def adjust_excel_column_widths( filename):
    # Load the workbook
    workbook = load_workbook(filename)

    for sheetname in workbook.sheetnames:
        worksheet = workbook[sheetname]
        if sheetname == "Summary":
            adjust_column_width(worksheet, multiplier=1)
        else:
            adjust_column_width(worksheet, multiplier=1)

    # Save the modified workbook back
    workbook.save(filename)

def adjust_column_width( worksheet, multiplier=1.1):
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length * multiplier
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

def adjust_column_widths_for_varied_sheets( filename, sheet_specs):
    # Load the workbook
    workbook = load_workbook(filename)

    # Iterate through each sheet
    for sheet_name in workbook.sheetnames:
        for phrase, column_widths in sheet_specs.items():
            if phrase in sheet_name:
                worksheet = workbook[sheet_name]

                # Adjust column widths
                for column_letter, width in column_widths.items():
                    worksheet.column_dimensions[column_letter].width = width
                break  # Stop checking other phrases once a match is found

    # Save the modified workbook back
    workbook.save(filename)

def clear_cells_fixed( filename):
    """
    Clear the values from cells O22 and O23 in an Excel file.

    Parameters:
    - filename: Path to the Excel file
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Assuming the modifications are for the first sheet (as per previous context)
    worksheet = workbook.active

    # Fixed cell addresses to clear
    cell_addresses = ["O22", "O23"]

    # Clear the values from specified cells
    for cell_address in cell_addresses:
        worksheet[cell_address].value = None

    # Save the modified workbook back to the same file
    workbook.save(filename)

def color_excel_tabs_inplace( filename):
    color_order = [
        "CCC0DA",
        "FFFF99",
        "00B0F0",
        "C4BD97",
        "CCC0DA",
        "DA9694",
        "E6B8B7",
        "FCD5B4",
        "C4D79B",
        "FCD5B4",
        "BFBFBF",
        "92D050",
    ]
    summary_sheet_color = "16365C"

    # Open the workbook
    workbook = openpyxl.load_workbook(filename)

    # Iterate through each sheet and set the tab color
    for idx, sheet_name in enumerate(workbook.sheetnames):
        if sheet_name == "Summary":
            workbook[sheet_name].sheet_properties.tabColor = summary_sheet_color
        else:
            color = color_order[idx % len(color_order)]
            workbook[sheet_name].sheet_properties.tabColor = color

    # Save the workbook
    workbook.save(filename)

def center_text_in_bankwise_sheets(
     filename, bankwise_sheet_prefix="BankWise Eligibility"
):
    wb = openpyxl.load_workbook(filename)

    # Create an Alignment style object for centering text
    center_aligned_text = Alignment(horizontal="center", vertical="center")

    # Define a border style for the bottom border
    bottom_border = Border(bottom=Side(border_style="thick", color="000000"))
    left_border = Border(left=Side(border_style="thick", color="000000"))
    left_bottom_border = Border(
        left=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000"),
    )

    # Loop through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        # Check if the sheet name contains the specified prefix
        if bankwise_sheet_prefix in sheet_name:
            ws = wb[sheet_name]

            # Apply center alignment to all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = center_aligned_text

            # Apply left border to cells from D1 to D18
            for cell in ws["D1:D18"]:
                cell[0].border = Border(
                    left=left_border.left
                )  # Apply only the left border

            for cell in ws["F1:F18"]:
                cell[0].border = Border(
                    left=left_border.left
                )  # Apply only the left border

            for cell in ws["H1:H18"]:
                cell[0].border = Border(
                    left=left_border.left
                )  # Apply only the left border

            for cell in ws["J1:J18"]:
                cell[0].border = Border(
                    left=left_border.left
                )  # Apply only the left border

            # Apply bottom borders to cells from D18 to I18
            for cell in ws["D18":"I18"][0]:  # Access the first row in the slice
                cell.border = bottom_border

            # Apply bottom borders to cells from D18 to I18
            for cell in ws["A18":"C18"][0]:  # Access the first row in the slice
                cell.border = bottom_border

            # Apply left and bottom borders to cells D18, F18, and H18
            for cell_ref in ["D18", "F18", "H18"]:
                ws[cell_ref].border = left_bottom_border

    # Save the workbook after making changes
    wb.save(filename)

def color_summary_sheet( filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb["Summary"]  # Access the "Summary" sheet

    # Initialize row variables
    row_6 = row_12 = row_30 = row_49 = row_66 = row_55 = row_56 = None

    # Search for the target words in the sheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "Particulars" in cell.value:
                    row_6 = cell.row
                elif "Income / Receipts" in cell.value:
                    row_12 = cell.row
                elif "Important Expenses / Payments" in cell.value:
                    row_30 = cell.row
                elif "Other Expenses / Payments" in cell.value:
                    row_49 = cell.row
                elif "Contra Credit" in cell.value:
                    row_55 = cell.row
                elif "Contra Debit" in cell.value:
                    row_56 = cell.row

    for cell in ws[1]:
        cell.value = None
        cell.border = Border()

    # Define colors
    fill_color = PatternFill(
        start_color="b9b9b9", end_color="b9b9b9", fill_type="solid"
    )
    royal_blue_fill = PatternFill(
        start_color="000058", end_color="000058", fill_type="solid"
    )
    white_bold_font = Font(bold=True, color="FFFFFF")
    light_blue_fill = PatternFill(
        start_color="b5cbe0", end_color="b5cbe0", fill_type="solid"
    )
    white_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    border_right = Border(
        right=Side(border_style="thin", color="000000")
    )  # thin black line
    border_thick_bottom = Border(
        bottom=Side(border_style="thin", color="000000")
    )  # thick black line

    for row in range(2, 5):  # Rows in openpyxl are 1-indexed
        for cell in ws[row]:
            cell.fill = fill_color
            cell.font = bold_font
    for row in [row_6, row_12, row_30, row_49, row_55, row_56]:
        for cell in ws[row]:
            cell.fill = royal_blue_fill
            cell.font = white_bold_font
    skip_rows = {row_6, row_12, row_30, row_49, row_55, row_56}

    for start_row in [row_6, row_12, row_30, row_49, row_55, row_56]:
        for row in range(start_row + 1, ws.max_row + 1):
            if row in skip_rows:
                continue  # Skip the rows that are in the skip list

            for cell in ws[row]:
                # Apply alternating color fill
                cell.fill = white_fill if (row - start_row) % 2 == 1 else light_blue_fill

    # for row_num in [16, 27, 41, 54]:
    #     for cell in ws[row_num]:
    #         cell.fill = white_fill
    for row in ws["A"]:
        row.border = border_right
    # for cell in ws[row_66]:
    #     cell.border = border_thick_bottom

    def apply_alternating_fill(sheet):

        for cell in sheet[1]:
            cell.fill = royal_blue_fill
            cell.font = white_bold_font

        for row in sheet.iter_rows(min_row=3):
            fill = light_blue_fill if (row[0].row - 3) % 2 == 0 else white_fill
            for cell in row:
                cell.fill = fill

    for sheet_name in wb.sheetnames[1:]:
        ws = wb[sheet_name]
        apply_alternating_fill(ws)

    # Save the changes
    wb.save(filename)

def color_summary_sheet_dsa( filename):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb["Summary"]  # Access the "Summary" sheet
    ws.sheet_properties.tabColor = "1072BA"
    for cell in ws[1]:
        cell.value = None
        cell.border = Border()

    # Define colors
    fill_color = PatternFill(
        start_color="b9b9b9", end_color="b9b9b9", fill_type="solid"
    )
    royal_blue_fill = PatternFill(
        start_color="000058", end_color="000058", fill_type="solid"
    )
    white_bold_font = Font(bold=True, color="FFFFFF")
    light_blue_fill = PatternFill(
        start_color="b5cbe0", end_color="b5cbe0", fill_type="solid"
    )
    white_fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )
    border_right = Border(
        right=Side(border_style="thin", color="000000")
    )  # thin black line
    border_thick_bottom = Border(
        bottom=Side(border_style="thin", color="000000")
    )  # thick black line

    for cell in ws[1]:
        cell.value = None
        cell.border = Border()

        # Applying grey fill and bold font to rows 2-4
    for row in range(2, 5):
        for cell in ws[row]:
            cell.fill = fill_color
            cell.font = bold_font

        # Applying royal blue fill and white bold font to row 6
    for cell in ws[6]:
        cell.fill = royal_blue_fill
        cell.font = white_bold_font

        # Applying alternate white and light blue fills from row 7 to row 23
    for row in range(7, 24):
        fill_to_use = light_blue_fill if row % 2 == 0 else white_fill
        for cell in ws[row]:
            cell.fill = fill_to_use

        # Applying right border to column A
    for cell in ws["A"]:
        cell.border = border_right

        # Applying thick bottom border to row 23
    for cell in ws[23]:
        cell.border = border_thick_bottom

    wb.save(filename)

    def apply_alternating_fill(sheet):

        for cell in sheet[1]:
            cell.fill = royal_blue_fill
            cell.font = white_bold_font

        for row in sheet.iter_rows(min_row=3):
            fill = light_blue_fill if (row[0].row - 3) % 2 == 0 else white_fill
            for cell in row:
                cell.fill = fill

    for sheet_name in wb.sheetnames[1:]:
        ws = wb[sheet_name]
        apply_alternating_fill(ws)

    # Save the changes
    wb.save(filename)

def format_numbers_with_commas( filename):
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Iterate through each sheet
    for sheet in workbook.worksheets:
        # Iterate through each row and column
        for row in sheet.iter_rows(
            min_row=2, max_col=sheet.max_column, max_row=sheet.max_row
        ):
            for cell in row:
                # Check if the cell contains a number
                if isinstance(cell.value, (int, float)):
                    # Set the number format
                    cell.number_format = "#,##0.00"

    # Save the workbook
    workbook.save(filename)

def add_filters_to_excel( filename):
    workbook = openpyxl.load_workbook(filename)

    for sheet in workbook.sheetnames[1:]:
        worksheet = workbook[sheet]
        worksheet.auto_filter.ref = worksheet.dimensions

    workbook.save(filename)

def approximate_width( cell_content):
    """
    Approximates the width of a cell's content.

    Parameters:
    cell_content (str): The content of the cell.

    Returns:
    float: An approximate width of the cell content.
    """
    max_length = 0
    for char in cell_content:
        if char.islower():
            max_length += 1.3  # Approx width for lowercase
        else:
            max_length += 1.3  # Approx width for uppercase and numbers
    return max_length

def auto_adjust_column_width( filename):
    """
    Automatically adjusts the column widths in all sheets of an Excel workbook based on the largest cell in each column.

    Parameters:
    filename (str): Path to the Excel file.
    """
    workbook = openpyxl.load_workbook(filename)

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        column_widths = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Approximate the width of the cell's content
                    estimated_width = approximate_width(str(cell.value))
                    column = cell.column_letter
                    column_widths[column] = max(
                        column_widths.get(column, 0), estimated_width
                    )

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

    workbook.save(filename)

def Eligibility_note(
     filename, empty_rows_between=1, sheet_name_keyword="BankWise Eligibility"
):
    note_parts = [
        "Disclaimer:",
        "1. The above loan eligibility calculations applies to self-employed clients and on the basis of Average Bank Balance only ",
        "2. The above eligibility is based on the analysis of the current uploaded bank statement. Kindly upload all bank statements to obtain more accurate eligibility.",
        "3. Final Approval will be dependent on complete thorough process and submission of relevant documents, CIBIL check, etc. by the respective Banks/NBFC's.",
        "4. Nothing contained in this eligibility should be deemed to create any right and/or interest whatsoever in favor of or against any party.",
        "5. Client data is not stored in our portal",
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        target_sheet = None

        # Find a sheet that contains the keyword
        for name in workbook.sheetnames:
            if sheet_name_keyword in name:
                target_sheet = workbook[name]
                break

        if not target_sheet:
            print(
                f"No sheet containing '{sheet_name_keyword}' found in the workbook."
            )
            return

        last_row = target_sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 5}"  # Extend the merge to 5 rows below the start
        )

        target_sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = target_sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def Summary_note( filename, empty_rows_between=1, sheet_name="Summary"):
    note_parts = [
        "Disclaimer/Caveat: The entries throughout this file and tables are based on best guess basis and",
        "information filtered under expenses and income. An attempt has been made to reflect the narration as",
        "close as possible to the actuals. However, variations from above are possible based on customer profile",
        "and their transactions with parties. Kindly cross check with your clients for any discrepancies",
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 4}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def Eod_note( filename, bank_names, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    for bank_name in bank_names:
        sheet_name = f"{bank_name} EOD Balance"
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            last_row = sheet.max_row
            sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
        else:
            print(f"No sheet named {sheet_name} found in {filename}.")
    workbook.save(filename)

def Comon_Eod_note( filename, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[f"Combined EOD Balance"]
    last_row = sheet.max_row
    sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
    workbook.save(filename)

def Investment_note( filename, empty_rows_between=1, sheet_name="Investment"):
    note_parts = [
        "*This table reflects probable transactions in securities made during the year. \r\nKindly confirm the same from Annual Information Statement (AIS) reflected on the Income Tax Portal and the capital gain reports sent by the respective authorities."
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def CreditorList_note( filename, empty_rows_between=1, sheet_name="Creditors"):
    note_parts = [
        "*The entries in this table likely pertain to payments from the parties during the period mentioned. \r\nIn case of payments through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a creditor."
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def DebtorList_note( filename, empty_rows_between=1, sheet_name="Debtors"):
    note_parts = [
        "*The entries in this table likely pertains to receipts from the respective parties. \r\nIn case of receipts through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a debtor."
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def CashWithdrawalt_note( filename, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Cash Withdrawal"]
    last_row = sheet.max_row
    sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
    workbook.save(filename)

def Cash_Deposit_note( filename, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Cash Deposit"]
    last_row = sheet.max_row
    sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
    workbook.save(filename)

def Emi_note(
     filename, note, empty_rows_between=1, column=1, sheet_name="Probable EMI"
):
    note = note.replace("\n", "\r\n")
    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1
        for i, line in enumerate(note.split("\r\n")):
            sheet.cell(row=start_row + i, column=column, value=line)
        workbook.save(filename)
    except:
        print("statement")

def create_excel_sheet( save_path, loan_value_df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Color

    # Include any other specific classes or functions you need from openpyxl

    # Check if DataFrame is empty or if specific columns are missing
    if (
        loan_value_df.empty
        or "Maximum Home Loan Value" not in loan_value_df.columns
        or "Maximum LAP Value" not in loan_value_df.columns
        or "Maximum BL Value" not in loan_value_df.columns
    ):
        print("DataFrame is empty or required columns are missing.")
        return  # or handle this situation in a way that fits your application

    # Your existing logic with safety checks
    max_home_loan = (
        0
        if pd.isna(loan_value_df["Maximum Home Loan Value"].iloc[0])
        else round(loan_value_df["Maximum Home Loan Value"].iloc[0] / 1000) * 1000
    )
    max_lap = (
        0
        if pd.isna(loan_value_df["Maximum LAP Value"].iloc[0])
        else round(loan_value_df["Maximum LAP Value"].iloc[0] / 1000) * 1000
    )
    max_bl = (
        0
        if pd.isna(loan_value_df["Maximum BL Value"].iloc[0])
        else round(loan_value_df["Maximum BL Value"].iloc[0] / 1000) * 1000
    )

    commission_percentage = [
        0.45 / 100,
        0.65 / 100,
        1.00 / 100,
    ]  # Convert percentages to fractions

    # Calculate commission in Rs
    commission_home_loan = round(max_home_loan * commission_percentage[0], 2)
    commission_lap = round(max_lap * commission_percentage[1], 2)
    commission_bl = round(max_bl * commission_percentage[2], 2)

    try:
        wb = load_workbook(save_path)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.create_sheet(title="Opportunity to Earn")
    content_font = Font(size=11)
    thin_side = Side(border_style="thin", color="000000")  # color is optional

    # Create the border using the defined sides
    border = Border(
        top=thin_side, left=thin_side, right=thin_side, bottom=thin_side
    )
    # Set column widths
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    # Define fonts and styles
    header_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")
    thin_border_side = Side(border_style="thin")
    header_border = Border(
        top=thin_border_side,
        bottom=thin_border_side,
        left=thin_border_side,
        right=thin_border_side,
    )

    # Add headers and apply styles
    headers = ["B4", "C4", "D4", "E4"]
    titles = ["Product", "Amount", "Commission %", "Commission (in Rs)"]

    for cell, title in zip(headers, titles):
        ws[cell] = title
        ws[cell].font = header_font
        ws[cell].alignment = center_aligned_text
        ws[cell].border = header_border

    # You may also want to set the width of the columns to make sure the headers fit properly
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 45

    # Create the dataframe
    data = {
        "Product": [
            "Home Loan / Balance Transfer",
            "Loan Against Property / Balance Transfer",
            "Business Loan",
            "Term Plan",
            "General Insurance",
        ],
        "Amount": [max_home_loan, max_lap, max_bl, np.nan, np.nan],
        "Commission %": ["0.45%", "0.65%", "1.00%", "1%-30%", "upto 10%"],
        "Commission (in Rs)": [
            commission_home_loan,
            commission_lap,
            commission_bl,
            np.nan,
            np.nan,
        ],
    }
    df = pd.DataFrame(data)

    # Add data to Excel from dataframe
    for r_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=False), 5
    ):
        for c_idx, value in enumerate(row, 2):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx).font = content_font
            ws.cell(row=r_idx, column=c_idx).border = border

    # Merge cells and add texts
    from openpyxl.styles import Alignment, Font
    from openpyxl.styles.colors import BLUE

    # Assuming 'ws' is your worksheet and 'header_font' is your predefined font
    ws.merge_cells("B2:E2")
    ws["B2"] = (
        "Congratulations!!! Here is an Opportunity for You to Earn.\n\nYour Client is Eligible for the following products"
    )

    # Define the font with the desired dark blue color. Excel uses ARGB format, 'FF' for solid.
    dark_blue_font = Font(
        color="FF134F63",
        size=header_font.size,
        bold=header_font.bold,
        italic=header_font.italic,
    )
    ws["B2"].alignment = Alignment(wrap_text=True)

    ws.row_dimensions[2].height = 60
    # Set the font to the cell
    ws["B2"].font = dark_blue_font

    # Define the alignment to left-aligned text.
    left_aligned_text = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    # Set the alignment to the cell
    ws["B2"].alignment = left_aligned_text

    title_text = "To Proceed Further:"
    title_cell = ws["B10"]
    title_cell.value = title_text
    title_cell.font = Font(bold=True, size=11)
    title_cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )

    # Add the detailed message to B11 without bold
    detail_message = (
        "***In case your client is interested in any of the above products, you can contact our trusted "
        "vendor M/s BizPedia Tech Private Limited on 8828824242 and email id support@leadsathi.in. "
        'Kindly use the promo code "CYPHERSOLEARN" to avail the higher commission structure.\n\n'
        "Once the referrals are successfully closed, you will be eligible for payouts based on the above commission structure.\n\n"
        "The respective payments will be released on the 25th of the next month."
    )

    detail_message_cell = ws["B11"]
    detail_message_cell.value = detail_message
    detail_message_cell.font = Font(size=11)
    detail_message_cell.alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    ws.row_dimensions[11].height = 100
    ws.merge_cells("B10:E10")
    ws.merge_cells("B11:E11")

    merge_ranges_to_unmerge = [
        str(merge_range)
        for merge_range in ws.merged_cells.ranges
        if "B13" in merge_range
    ]

    # Unmerge them in a separate loop
    for merge_range in merge_ranges_to_unmerge:
        ws.unmerge_cells(merge_range)

    # Now you can safely set the value of the top-left cell of the range you intend to merge
    ws["B13"] = (
        "Disclaimer:\n\n1. The above loan eligibility calculations apply to self-employed clients only.\n"
        "2. For salaried clients, the vendor will need more details to calculate the eligibility.\n"
        "3. The above eligibility is based on the analysis of the current uploaded bank statement. Kindly upload "
        "all bank statements to obtain more accurate eligibility.\n"
        "4. Final Approval will be dependent on complete thorough process and submission of relevant documents, "
        "CIBIL check, etc.\n"
        "5. Nothing contained in this eligibility should be deemed to create any right and/or interest whatsoever "
        "in favor of or against any party."
    )

    # Set the alignment and font for the cell before merging
    cell = ws["B13"]
    cell.alignment = Alignment(wrap_text=True)
    cell.font = Font(size=11)

    # Merge the cells after setting the value
    ws.merge_cells("B13:E21")

    for row in ws.iter_rows(min_row=4, max_row=9, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = center_aligned_text
    # Save the workbook
    wb.save(save_path)

    return save_path

def Entertainment_note( filename, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Subscription_Entertainment"]
    last_row = sheet.max_row
    sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
    workbook.save(filename)

def Refund_note( filename, note, empty_rows_between=1, column=1):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Refund-Reversal"]
    last_row = sheet.max_row
    sheet.cell(row=last_row + empty_rows_between, column=column, value=note)
    workbook.save(filename)

def Suspense_Credit_note(
     filename, empty_rows_between=1, sheet_name="Suspense Credit"
):
    note_parts = [
        "*This table pertains to transactions unidentified as per the current ledger bifurcation of the software. \r\nIn case of any technical errors, inconvience is highly regretted and feedback is appreciated."
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def Suspense_Debit_note(
     filename, empty_rows_between=1, sheet_name="Suspense Debit"
):
    note_parts = [
        "*This table likely pertains to transactions unidentified as per the current ledger bifurcation of the software."
    ]
    note = "\r\n".join(note_parts)

    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name not in workbook.sheetnames:
            print(f"The sheet {sheet_name} does not exist in the workbook.")
            return
        sheet = workbook[sheet_name]
        last_row = sheet.max_row
        start_row = last_row + empty_rows_between if last_row else 1

        merge_start = f"A{start_row}"
        merge_end = (
            f"H{start_row + 2}"  # Extend the merge to 3 rows below the start
        )

        sheet.merge_cells(f"{merge_start}:{merge_end}")

        # Add the text to the merged cells and enable text wrapping
        cell = sheet[merge_start]
        cell.value = note
        cell.alignment = Alignment(wrap_text=True)  # Enable wrap_text for the cell

        workbook.save(filename)
    except Exception as e:
        print(f"An error occurred: {e}")

def process_excel_to_json(excel_file):
        output_file = "excel_to_json.json"
        """
        Processes an Excel workbook and converts it to JSON format.
        """

        # Function to make column names unique globally
        def make_columns_unique(columns):
            seen = set()
            unique_columns = []
            for col in columns:
                if col in seen:
                    count = 1
                    new_col = f"{col}_{count}"
                    while new_col in seen:
                        count += 1
                        new_col = f"{col}_{count}"
                    unique_columns.append(new_col)
                    seen.add(new_col)
                else:
                    unique_columns.append(col)
                    seen.add(col)
            return unique_columns

        # Function to process the "Summary" sheet
        def summary_to_json(df):
            tables = {}
            current_table_data = []
            table_count = 0

            # Iterate through rows to extract tables
            for index, row in df.iterrows():
                if row.isnull().all():  # Blank row indicates the end of a table
                    if current_table_data:  # Process the current table if data exists
                        table_count += 1
                        table_df = pd.DataFrame(current_table_data)
                        table_df.columns = make_columns_unique(table_df.iloc[0])  # Use the first row as headers
                        table_df = table_df[1:]  # Remove the headers from the data
                        table_df.reset_index(drop=True, inplace=True)
                        tables[f"Table {table_count}"] = json.loads(table_df.to_json(orient="records"))
                        current_table_data = []  # Reset for the next table
                else:
                    current_table_data.append(row)

            # Process the last table if it exists
            if current_table_data:
                table_count += 1
                table_df = pd.DataFrame(current_table_data)
                table_df.columns = make_columns_unique(table_df.iloc[0])  # Use the first row as headers
                table_df = table_df[1:]  # Remove the headers from the data
                table_df.reset_index(drop=True, inplace=True)
                tables[f"Table {table_count}"] = json.loads(table_df.to_json(orient="records"))

            return tables

        """ 
        Args:
            excel_file (str): Path to the Excel workbook.
            output_file (str): Path to save the output JSON file.
        """
        excel_data = pd.ExcelFile(excel_file)
        result = {}

        # Process each sheet
        for sheet_name in excel_data.sheet_names:
            df = excel_data.parse(sheet_name, header=None)  # Load without headers

            if sheet_name == "Summary":  # Special case for the "Summary" sheet
                result[sheet_name] = summary_to_json(df)
            else:  # Process other sheets
                tables = []
                headers = df.iloc[0]  # First row is the header
                df = df[1:]  # Remove the header row from the data
                headers = headers.fillna(f"Unnamed_{len(headers)}")  # Fill NaNs in headers
                df.columns = make_columns_unique(headers)  # Ensure unique column names
                tables.append(df)

                # Process tables and add to JSON with table names
                sheet_result = {}
                for index, table in enumerate(tables):
                    table.reset_index(drop=True, inplace=True)
                    index_plus_one = f"Table {index + 1}"
                    sheet_result[index_plus_one] = json.loads(table.to_json(orient='records'))

                result[sheet_name] = sheet_result

        # Convert the result dictionary to JSON
        final_json = json.dumps(result, indent=4)

        print(f"JSON output generated")

        return final_json


def sort_dataframes_by_date(dataframes):
    """
    Sorts a list of dataframes based on the range of the 'Value Date' column.

    Args:
        dataframes (list of pd.DataFrame): List of dataframes with a common 'Value Date' column.

    Returns:
        list of pd.DataFrame: List of dataframes sorted by the earliest 'Value Date' in ascending order.
    """
    # Extract ranges and sort the dataframes
    ranges = []

    for i, df in enumerate(dataframes):
        if type(df) != pd.DataFrame:
            continue
        # Ensure 'Value Date' is of datetime type
        df['Value Date'] = pd.to_datetime(df['Value Date'], format='%d-%m-%Y', errors='coerce')

        # Get the earliest and latest dates
        start_date = df['Value Date'].min()
        end_date = df['Value Date'].max()

        # Store the index and range
        ranges.append((i, start_date, end_date))


    # Sort ranges by start_date
    sorted_ranges = sorted(ranges, key=lambda x: x[1])

    # Sort the dataframes based on the sorted indices
    sorted_dataframes = [dataframes[i] for i, _, _ in sorted_ranges]

    return sorted_dataframes


def get_missing_months(opening_closing_balance, balances):

    # Extract months from both dictionaries
    opening_closing_months = set(opening_closing_balance.keys())
    balances_months = set(balances.keys())

    # Find months in balances but not in opening_closing_balance
    missing_months = list(balances_months - opening_closing_months)

    missing_months.sort(key=lambda x: pd.to_datetime(x, format='%b-%Y'))

    return missing_months


def process_transactions(df):


    # Define the columns for the output
    columns = [
        "Date",
        "Effective Date",
        "Bill Ref",
        "Dr Ledger",
        "Cr Ledger",
        "Amount",
        "Voucher Type",
        "Narration",
    ]

    # Create an empty output DataFrame with the desired columns
    out_df = pd.DataFrame(columns=columns)

    # Iterate over each row in the original DataFrame
    for _, row in df.iterrows():
        # Prepare a basic dictionary for the new row
        new_row = {
            "Date": row["Value Date"],
            "Effective Date": "",  # fill if needed
            "Bill Ref": "-",  # fill if needed
            "Dr Ledger": "",
            "Cr Ledger": "",
            "Amount": 0,
            "Voucher Type":"",
            "Narration": row["Description"],
        }

        # Decide if this row is a Debit or a Credit
        if pd.notna(row.get("Debit", None)):
            # It's a payment (Debit)
            new_row["Dr Ledger"] = row["Category"]
            new_row["Voucher Type"] = "Payment"
            new_row["Amount"] = row["Debit"]
        elif pd.notna(row.get("Credit", None)):
            # It's a receipt (Credit)
            new_row["Cr Ledger"] = row["Category"]
            new_row["Voucher Type"] = "Receipt"
            new_row["Amount"] = row["Credit"]
        else:
            # If both are null or not as expected, handle however you wish
            # e.g., skip, or continue
            continue

        # Append the new_row to out_df
        out_df = out_df._append(new_row, ignore_index=True)

    return out_df


def get_total_pdf_pages(pdf_paths):
    """
    Calculate the total number of pages from a list of PDF file paths.

    :param pdf_paths: List of PDF file paths
    :return: Total number of pages
    """
    total_pages = 0

    for pdf_path in pdf_paths:
        if pdf_path:  # Ensure the path is not empty
            try:
                with fitz.open(pdf_path) as doc:
                    total_pages += len(doc)
            except Exception as e:
                print(f"Error processing {pdf_path}: {e}")

    return total_pages