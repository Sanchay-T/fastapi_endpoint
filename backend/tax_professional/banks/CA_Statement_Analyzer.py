import shutil
import os
from lib2to3.pytree import convert
from openpyxl.styles import Font
import logging
from openpyxl import Workbook, load_workbook
import sys
import json
import pandas as pd
import regex as re
import fitz
import os


bold_font = Font(bold=True)
pd.options.display.float_format = "{:,.2f}".format
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)
BASE_DIR = os.path.dirname(os.path.abspath(os.path.join(__file__, "../../../")))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../..")))
from ...utils import get_saved_pdf_dir, get_saved_excel_dir

# from findaddy.exceptions import ExtractionError
TEMP_SAVED_PDF_DIR = get_saved_pdf_dir()
TEMP_SAVED_EXCEL_DIR = get_saved_excel_dir()

from ...common_functions import (process_excel_to_json, process_name_n_num_df, category_add_ca,
                                 another_method, Upi, eod, opening_and_closing_bal, summary_sheet,
                                 transaction_sheet, total_investment, redemption_investment,
                                 creditor_list, debtor_list, cash_withdraw, cash_depo, div_int, emi,
                                 refund_reversal, suspense_credit, suspense_debit, payment, receipt,
                                 calculate_fixed_day_average, process_avg_last_6_months, extraction_process,
                                 color_summary_sheet, format_numbers_with_commas,
                                 adjust_column_widths_for_varied_sheets,
                                 Summary_note, Investment_note, CreditorList_note, DebtorList_note,
                                 CashWithdrawalt_note,
                                 Cash_Deposit_note, Emi_note, Refund_note, Suspense_Credit_note, Suspense_Debit_note,
                                 add_filters_to_excel, create_excel_sheet, color_excel_tabs_inplace,
                                 sort_dataframes_by_date,
                                 extraction_process_explicit_lines, process_transactions, get_total_pdf_pages)


def save_to_excel(df, name_n_num_df, account_number):
    # Generate all necessary DataFrames
    eod_sheet_df = eod(df)
    opening_bal, closing_bal = opening_and_closing_bal(eod_sheet_df)

    summary_df_list, missing_months_list = summary_sheet(df, opening_bal, closing_bal, df)

    particulars_df = summary_df_list[0]
    income_receipts_df = summary_df_list[1]
    imp_expenses_payments_df = summary_df_list[2]
    other_expenses_df = summary_df_list[3]
    contra_credit_df = summary_df_list[4]
    contra_debit_df = summary_df_list[5]

    df['Value Date'] = pd.to_datetime(df['Value Date'], format='%d-%m-%Y').dt.strftime('%d-%m-%Y')
    transaction_sheet_df = transaction_sheet(df)
    investment_df = total_investment(df)
    creditor_df = creditor_list(df)
    debtor_df = debtor_list(transaction_sheet_df)

    upi_cr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Credit"] > 0)]
    upi_dr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Debit"] > 0)]

    cash_withdrawal_df = cash_withdraw(df)
    cash_deposit_df = cash_depo(df)
    dividend_int_df = div_int(df)
    emi_df = emi(df)
    refund_df = refund_reversal(df)
    suspense_credit_df = suspense_credit(df)
    suspense_debit_df = suspense_debit(df)
    # payment_df = payment(df)
    # receipt_df = receipt(df)
    pay_n_receipt_df = process_transactions(df)

    bank_avg_balance_df = calculate_fixed_day_average(eod_sheet_df)
    loan_value_df = process_avg_last_6_months(bank_avg_balance_df, eod_sheet_df)

    os.makedirs(TEMP_SAVED_EXCEL_DIR, exist_ok=True)
    filename = os.path.join(TEMP_SAVED_EXCEL_DIR, f"Bank_{account_number}_Extracted_statements_file.xlsx")

    writer = pd.ExcelWriter(filename, engine="xlsxwriter")

    # start converting to excel
    sheet_name = "Summary"
    name_n_num_df.to_excel(writer, sheet_name=sheet_name, index=False)

    particulars_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0] + 2,
        index=False,
    )
    income_receipts_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0] + particulars_df.shape[0] + 4,
        index=False,
    )
    imp_expenses_payments_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0]
                 + particulars_df.shape[0]
                 + income_receipts_df.shape[0]
                 + 6,
        index=False,
    )
    other_expenses_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0]
                 + particulars_df.shape[0]
                 + income_receipts_df.shape[0]
                 + imp_expenses_payments_df.shape[0]
                 + 8,
        index=False,
    )

    contra_credit_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0]
                 + particulars_df.shape[0]
                 + income_receipts_df.shape[0]
                 + imp_expenses_payments_df.shape[0]
                 + other_expenses_df.shape[0]
                 + 10,
        index=False,
    )

    contra_debit_df.to_excel(
        writer,
        sheet_name=sheet_name,
        startrow=name_n_num_df.shape[0]
                 + particulars_df.shape[0]
                 + income_receipts_df.shape[0]
                 + imp_expenses_payments_df.shape[0]
                 + other_expenses_df.shape[0]
                 + contra_credit_df.shape[0]
                 + 12,
        index=False,
    )

    loan_value_df.to_excel(writer, sheet_name='Opportunity to Earn', index=False)
    transaction_sheet_df.to_excel(writer, sheet_name='Transactions', index=False)
    eod_sheet_df.to_excel(writer, sheet_name='EOD', index=False)
    investment_df.to_excel(writer, sheet_name='Investment', index=False)
    creditor_df.to_excel(writer, sheet_name='Creditors', index=False)
    debtor_df.to_excel(writer, sheet_name='Debtors', index=False)
    upi_cr_df.to_excel(writer, sheet_name='UPI-CR', index=False)
    upi_dr_df.to_excel(writer, sheet_name='UPI-DR', index=False)
    cash_withdrawal_df.to_excel(writer, sheet_name='Cash Withdrawal', index=False)
    cash_deposit_df.to_excel(writer, sheet_name='Cash Deposit', index=False)
    dividend_int_df.to_excel(writer, sheet_name='Redemption, Dividend & Interest', index=False)
    emi_df.to_excel(writer, sheet_name='Probable EMI', index=False)
    refund_df.to_excel(writer, sheet_name='Refund-Reversal', index=False)
    suspense_credit_df.to_excel(writer, sheet_name='Suspense Credit', index=False)
    suspense_debit_df.to_excel(writer, sheet_name='Suspense Debit', index=False)
    pay_n_receipt_df.to_excel(writer, sheet_name='Payment & Receipt Voucher', index=False)
    # payment_df.to_excel(writer, sheet_name='Payment Voucher', index=False)
    # receipt_df.to_excel(writer, sheet_name='Receipt Voucher', index=False)
    writer.close()

    # added color formatting nd filter functionalities
    color_summary_sheet(filename)
    format_numbers_with_commas(filename)
    sheet_specs = {
        "Summary": {
            "A": 50,
            "B": 15,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 15,
        },
        "DateWise Avg Balance": {
            "A": 35,
            "B": 15,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 20,
            "O": 20,
            "P": 20,
        },
        "BankWise Eligibility": {
            "A": 35,
            "B": 20,
            "C": 20,
            "D": 25,
            "E": 25,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 25,
        },
        "Transaction": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "EOD Balance": {
            "A": 15,
            "B": 15,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 15,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 15,
            "P": 15,
            "Q": 15,
            "R": 15,
            "S": 15,
            "T": 15,
            "U": 15,
            "V": 15,
            "W": 15,
            "X": 15,
            "Y": 15,
            "Z": 15,
        },
        "Probable EMI": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Bounce": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
        "Creditor": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Debtor": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Cash Deposit": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Cash Withdrawal": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "POS_CR": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
        "UPI_CR": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
        "Investment": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Subscription_Entertainment": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Refund-Reversal": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Suspense Credit": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Suspense Debit": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Redemption, Dividend & Interest": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Bank_charges": {
            "A": 10,
            "B": 70,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 20,
            "G": 10,
        },
        "Payment Voucher": {
            "A": 15,
            "B": 15,
            "C": 10,
            "D": 20,
            "E": 15,
            "F": 15,
            "G": 85,
        },
        "Receipt Voucher": {
            "A": 15,
            "B": 15,
            "C": 20,
            "D": 10,
            "E": 15,
            "F": 85,

        },
        "Payment_Receipt Voucher": {
            "A": 15,
            "B": 15,
            "C": 10,
            "D": 20,
            "E": 20,
            "F": 20,
            "G": 20,
            "H": 85,
        },
    }
    adjust_column_widths_for_varied_sheets(filename, sheet_specs)
    Summary_note(filename, empty_rows_between=2)
    Investment_note(filename, empty_rows_between=2)
    CreditorList_note(filename, empty_rows_between=2)
    DebtorList_note(filename, empty_rows_between=2)
    CW_note = "*The above table reflects the cash withdrawals made during the year on the basis of widely used acronyms of the finance industry."
    CashWithdrawalt_note(filename, CW_note, empty_rows_between=2)
    CD_note = "*The above table reflects the cash deposits made during the year on the basis of widely used acronyms of the finance industry."
    Cash_Deposit_note(filename, CD_note, empty_rows_between=2)
    emi_note = "* Transactions in the above table are based on the widely used acronyms of the finance industry and likely reflect EMI payment. \r\nKindly confirm the same from the loan statement or the interest certificate."
    Emi_note(filename, emi_note, empty_rows_between=2)
    reef_note = "*This table likely pertains to refunds/reversals/cashbacks received from card payments/online transactions."
    Refund_note(filename, reef_note, empty_rows_between=2)
    Suspense_Credit_note(filename, empty_rows_between=2)
    Suspense_Debit_note(filename, empty_rows_between=2)
    add_filters_to_excel(filename)
    create_excel_sheet(filename, loan_value_df)
    color_excel_tabs_inplace(filename)

    def reorder_sheets(filename):
        wb = load_workbook(filename)
        desired_order_front = ["Summary", "Opportunity to Earn"]
        existing_sheets = [
            sheet for sheet in desired_order_front if sheet in wb.sheetnames
        ]
        other_sheets = [
            sheet for sheet in wb.sheetnames if sheet not in existing_sheets
        ]
        new_order = existing_sheets + other_sheets
        wb._sheets = [wb[sheet] for sheet in new_order]
        wb.save(filename)
        return "Sheets reordered successfully"

    reorder_sheets(filename)

    return filename


def returns_json_output_of_all_sheets(df, name_n_num_df):
    # Generate all necessary DataFrames
    eod_sheet_df = eod(df)
    opening_bal, closing_bal = opening_and_closing_bal(eod_sheet_df)

    summary_df_list, missing_months_list = summary_sheet(df, opening_bal, closing_bal, df)

    particulars_df = summary_df_list[0]
    income_receipts_df = summary_df_list[1]
    imp_expenses_payments_df = summary_df_list[2]
    other_expenses_df = summary_df_list[3]
    contra_credit_df = summary_df_list[4]
    contra_debit_df = summary_df_list[5]

    df['Value Date'] = pd.to_datetime(df['Value Date']).dt.strftime('%d-%m-%Y')
    transaction_sheet_df = transaction_sheet(df)
    investment_df = total_investment(df)
    creditor_df = creditor_list(df)
    debtor_df = debtor_list(transaction_sheet_df)

    upi_cr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Credit"] > 0)]
    upi_dr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Debit"] > 0)]

    cash_withdrawal_df = cash_withdraw(df)
    cash_deposit_df = cash_depo(df)
    dividend_int_df = div_int(df)
    emi_df = emi(df)
    refund_df = refund_reversal(df)
    suspense_credit_df = suspense_credit(df)
    suspense_debit_df = suspense_debit(df)
    payment_df = payment(df)
    receipt_df = receipt(df)
    pay_n_receipt_df = process_transactions(df)

    bank_avg_balance_df = calculate_fixed_day_average(eod_sheet_df)
    loan_value_df = process_avg_last_6_months(bank_avg_balance_df, eod_sheet_df)

    # Build a dictionary to hold the labeled DataFrames
    result_dict = {
        "Name Acc No": name_n_num_df.to_dict(orient="records"),
        "Particulars": particulars_df.to_dict(orient="records"),
        "Income Receipts": income_receipts_df.to_dict(orient="records"),
        "Important Expenses": imp_expenses_payments_df.to_dict(orient="records"),
        "Other Expenses": other_expenses_df.to_dict(orient="records"),
        "Contra Credit": contra_credit_df.to_dict(orient="records"),
        "Contra Debit": contra_debit_df.to_dict(orient="records"),
        "Opportunity to Earn": loan_value_df.to_dict(orient="records"),
        "Transactions": transaction_sheet_df.to_dict(orient="records"),
        "EOD": eod_sheet_df.to_dict(orient="records"),
        "Investment": investment_df.to_dict(orient="records"),
        "Creditors": creditor_df.to_dict(orient="records"),
        "Debtors": debtor_df.to_dict(orient="records"),
        "UPI-CR": upi_cr_df.to_dict(orient="records"),
        "UPI-DR": upi_dr_df.to_dict(orient="records"),
        "Cash Withdrawal": cash_withdrawal_df.to_dict(orient="records"),
        "Cash Deposit": cash_deposit_df.to_dict(orient="records"),
        "Redemption, Dividend & Interest": dividend_int_df.to_dict(orient="records"),
        "Probable EMI": emi_df.to_dict(orient="records"),
        "Refund-Reversal": refund_df.to_dict(orient="records"),
        "Suspense Credit": suspense_credit_df.to_dict(orient="records"),
        "Suspense Debit": suspense_debit_df.to_dict(orient="records"),
        "Payment & Receipt Voucher": pay_n_receipt_df.to_dict(orient="records"),
        "Payment Voucher": payment_df.to_dict(orient="records"),
        "Receipt Voucher": receipt_df.to_dict(orient="records"),
    }

    # Convert the entire dictionary to JSON
    json_output = json.dumps(result_dict, indent=4)
    with open("new_output.json", "w") as file:
        file.write(json_output)
    return json_output, missing_months_list


def refresh_category_all_sheets(df,eod_sheet_df, new_categories):
    # eod_sheet_df = eod(df)
    opening_bal, closing_bal = opening_and_closing_bal(eod_sheet_df)

    if not new_categories:
        summary_df_list, missing_months_list = summary_sheet(df, opening_bal, closing_bal, df)
    else:
        summary_df_list, missing_months_list = summary_sheet(df, opening_bal, closing_bal, df, new_categories)

    particulars_df = summary_df_list[0]
    income_receipts_df = summary_df_list[1]
    imp_expenses_payments_df = summary_df_list[2]
    other_expenses_df = summary_df_list[3]
    contra_credit_df = summary_df_list[4]
    contra_debit_df = summary_df_list[5]

    df['Value Date'] = pd.to_datetime(df['Value Date']).dt.strftime('%d-%m-%Y')
    # transaction_sheet_df = transaction_sheet(df)
    # investment_df = total_investment(df)
    # creditor_df = creditor_list(df)
    # debtor_df = debtor_list(transaction_sheet_df)

    # upi_cr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Credit"] > 0)]
    # upi_dr_df = df[(df["Description"].str.contains("UPI", case=False)) & (df["Debit"] > 0)]

    # cash_withdrawal_df = cash_withdraw(df)
    # cash_deposit_df = cash_depo(df)
    # dividend_int_df = div_int(df)
    # emi_df = emi(df)
    # refund_df = refund_reversal(df)
    # suspense_credit_df = suspense_credit(df)
    # suspense_debit_df = suspense_debit(df)
    # payment_df = payment(df)
    # receipt_df = receipt(df)

    bank_avg_balance_df = calculate_fixed_day_average(eod_sheet_df)
    loan_value_df = process_avg_last_6_months(bank_avg_balance_df, eod_sheet_df)

    # Build a dictionary to hold the labeled DataFrames
    result_dict = {
        # "Name Acc No": name_n_num_df.to_dict(orient="records"),
        "Particulars": particulars_df.to_dict(orient="records"),
        "Income Receipts": income_receipts_df.to_dict(orient="records"),
        "Important Expenses": imp_expenses_payments_df.to_dict(orient="records"),
        "Other Expenses": other_expenses_df.to_dict(orient="records"),
        "Contra Credit": contra_credit_df.to_dict(orient="records"),
        "Contra Debit": contra_debit_df.to_dict(orient="records"),
        "Opportunity to Earn": loan_value_df.to_dict(orient="records"),
        # "Transactions": transaction_sheet_df.to_dict(orient="records"),
        # "EOD": eod_sheet_df.to_dict(orient="records"),
        # "Investment": investment_df.to_dict(orient="records"),
        # "Creditors": creditor_df.to_dict(orient="records"),
        # "Debtors": debtor_df.to_dict(orient="records"),
        # "UPI-CR": upi_cr_df.to_dict(orient="records"),
        # "UPI-DR": upi_dr_df.to_dict(orient="records"),
        # "Cash Withdrawal": cash_withdrawal_df.to_dict(orient="records"),
        # "Cash Deposit": cash_deposit_df.to_dict(orient="records"),
        # "Redemption, Dividend & Interest": dividend_int_df.to_dict(orient="records"),
        # "Probable EMI": emi_df.to_dict(orient="records"),
        # "Refund-Reversal": refund_df.to_dict(orient="records"),
        # "Suspense Credit": suspense_credit_df.to_dict(orient="records"),
        # "Suspense Debit": suspense_debit_df.to_dict(orient="records"),
        # "Payment Voucher": payment_df.to_dict(orient="records"),
        # "Receipt Voucher": receipt_df.to_dict(orient="records"),
    }

    # Convert the entire dictionary to JSON
    json_output = json.dumps(result_dict, indent=4)

    return json_output


def individual_summary(transactions_df):
    
    transactions_df.rename(columns={
        "description": "Description",
        "category": "Category",
        "entity": "Entity",
        "voucher_type": "Voucher type"
    }, inplace=True)
    
    eod_sheet_df = eod(transactions_df)
    print(eod_sheet_df.head(10))
    opening_bal, closing_bal = opening_and_closing_bal(eod_sheet_df)
    # named print 
    print("opening_bal", opening_bal)
    print("closing_bal", closing_bal)


    summary_df_list,mission_months = summary_sheet(transactions_df, opening_bal, closing_bal, transactions_df)

    print("summary_df_list", len(summary_df_list))
    particulars_df = summary_df_list[0]
    print("particulars_df", particulars_df)
    income_receipts_df = summary_df_list[1]
    print("income_receipts_df", income_receipts_df)
    imp_expenses_payments_df = summary_df_list[2]
    print("imp_expenses_payments_df", imp_expenses_payments_df)
    other_expenses_df = summary_df_list[3]
    print("other_expenses_df", other_expenses_df)
    contra_credit_df = summary_df_list[4]
    print("contra_credit_df", contra_credit_df)
    contra_debit_df = summary_df_list[5]
    print("contra_debit_df", contra_debit_df)

    result_dict = {
        "Particulars": particulars_df.to_dict(orient="records"),
        "Income Receipts": income_receipts_df.to_dict(orient="records"),
        "Important Expenses": imp_expenses_payments_df.to_dict(orient="records"),
        "Other Expenses": other_expenses_df.to_dict(orient="records"),
        "Contra Credit": contra_credit_df.to_dict(orient="records"),
        "Contra Debit": contra_debit_df.to_dict(orient="records"),
    }

    json_output = json.dumps(result_dict, indent=4)

    return json_output
    

def start_extraction_add_pdf(bank_names, pdf_paths, passwords, start_dates, end_dates, CA_ID, progress_data,
                             whole_transaction_sheet=None, aiyazs_array_of_array=None):
    account_number = ""
    dfs = {}
    name_dfs = {}
    errorz = {}

    pdf_paths_not_extracted = {
        "bank_names": [],
        "paths": [],
        "passwords": [],
        "start_dates": [],
        "end_dates": [],
        "respective_list_of_columns": [],
        "respective_reasons_for_error": []
    }
    i = 0

    for bank in bank_names:
        bank = str(f"{bank}{i}")
        pdf_path = pdf_paths[i]
        pdf_password = passwords[i]
        start_date = start_dates[i]
        end_date = end_dates[i]

        if aiyazs_array_of_array:
            aiyaz_array_of_array = aiyazs_array_of_array[i]
            print("aiyaz_array_of_array from ca statement analyzer - ", aiyaz_array_of_array)
            # Iterate through the columns to extract start and end coordinates
            explicit_lines = list(
                {coord for item in aiyaz_array_of_array for coord in (item["bounds"]["start"], item["bounds"]["end"])})
            labels = [[entry["index"], entry["column_type"]] for entry in aiyaz_array_of_array]
            dfs[bank], name_dfs[bank], errorz[bank] = extraction_process_explicit_lines(bank, pdf_path, pdf_password,
                                                                                        start_date, end_date,
                                                                                        explicit_lines, labels)

        else:
            dfs[bank], name_dfs[bank], errorz[bank] = extraction_process(bank, pdf_path, pdf_password, start_date,
                                                                         end_date)

        print(f"Extracted {bank} bank statement successfully")
        # account_number += f"{name_dfs[bank][1][:4]}x{name_dfs[bank][1][-4:]}_"
        # Check if the extracted dataframe is empty
        if dfs[bank].empty:
            pdf_paths_not_extracted["bank_names"].append(re.sub(r"\d+", "", bank))

            pdf_document = fitz.open(pdf_path)
            if pdf_document.is_encrypted:
                if not pdf_document.authenticate(pdf_password):
                    raise ValueError("Incorrect password. Unable to unlock the PDF.")

                # Save the unlocked PDF in the same location, replacing the original file
                temp_path = pdf_path + ".unlocked.pdf"
                pdf_document.save(temp_path)
                pdf_document.close()

                # Replace original file
                shutil.move(temp_path, pdf_path)

                print("PDF unlocked and saved successfully in the same path")

            pdf_paths_not_extracted["paths"].append(pdf_path)
            pdf_paths_not_extracted["passwords"].append(pdf_password)
            pdf_paths_not_extracted["start_dates"].append(start_date)
            pdf_paths_not_extracted["end_dates"].append(end_date)
            pdf_paths_not_extracted["respective_list_of_columns"].append(name_dfs[bank])
            pdf_paths_not_extracted["respective_reasons_for_error"].append(errorz[bank])
            del dfs[bank]
            del name_dfs[bank]

        i += 1

    print("|------------------------------|")
    print(account_number)
    print("|------------------------------|")

    if not dfs:
        folder_path = "saved_pdf"
        try:
            shutil.rmtree(folder_path)
            print(f"Removed all contents in '{folder_path}'")
        except Exception as e:
            print(f"Failed to remove '{folder_path}': {e}")

        return {"sheets_in_json": None, 'pdf_paths_not_extracted': pdf_paths_not_extracted, 'success_page_number': 0,
                'missing_months_list': []}

    else:
        data = []
        # num_pairs = len(pd.Series(dfs).to_dict())

        for key, value in name_dfs.items():
            bank_name = key
            acc_name = value[0]
            acc_num = value[1]
            if str(acc_num) == "None":
                masked_acc_num = "None"
            else:
                masked_acc_num = "X" * (len(acc_num) - 4) + acc_num[-4:]
            data.append([masked_acc_num, acc_name, bank_name])
            for item in data:
                item[2] = "".join(
                    character for character in item[2] if character.isalpha()
                )

        name_n_num_df = process_name_n_num_df(data)
        list_of_dataframes = list(dfs.values())

        if whole_transaction_sheet is not None:
            list_of_dataframes.append(whole_transaction_sheet)

        # arrange dfs
        initial_df = pd.concat(sort_dataframes_by_date(list_of_dataframes)).fillna("").reset_index(drop=True)
        initial_df = initial_df.drop_duplicates(keep="first")
        df = category_add_ca(initial_df)
        new_tran_df = another_method(df)
        new_tran_df = Upi(new_tran_df)
        # print("transaction")
        # print(new_tran_df)
        #############################------------------------#######################################

        json_lists_of_df, missing_months_list = returns_json_output_of_all_sheets(new_tran_df, name_n_num_df)

        # excel_file_path = save_to_excel(new_tran_df, name_n_num_df, account_number)
        # print(excel_file_path)

        all_pdf_pages = get_total_pdf_pages(pdf_paths)
        not_extracted_pages = get_total_pdf_pages(pdf_paths_not_extracted['paths'])
        time_saved_pages = all_pdf_pages - not_extracted_pages

        folder_path = "saved_pdf"
        try:
            shutil.rmtree(folder_path)
            print(f"Removed all contents in '{folder_path}'")
        except Exception as e:
            print(f"Failed to remove '{folder_path}': {e}")
        # print(name_n_num_df)

        print("%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%")
        print(pdf_paths_not_extracted)
        print("%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%")

        return {"sheets_in_json": json_lists_of_df, 'pdf_paths_not_extracted': pdf_paths_not_extracted,
                'success_page_number': time_saved_pages, 'missing_months_list': missing_months_list}

# # #
# bank_names = ["ICICI"]
# pdf_paths = ["F.Y. 2021-2022.pdf"]
# passwords = ["059501501877"]
# start_dates = ["01-09-2020"]
# end_dates = ["03-03-2025"]
# CA_ID = "CA_ID_4321"
# progress_data = {}
# null = "null"
#
# # # # x = start_extraction_edit_pdf(bank_names, pdf_paths, passwords, start_dates, end_dates, CA_ID, progress_data, aiyaz_array_of_array, whole_transaction_sheet=None)
# result = start_extraction_add_pdf(bank_names, pdf_paths, passwords, start_dates, end_dates, CA_ID, progress_data)
# print("exit")
# print(result["pdf_paths_not_extracted"])
